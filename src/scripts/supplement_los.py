from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from src.utils.db_utils import update_appflag, update_user, check_appflag, update_applicant
from src.utils.login_utils import perform_login
from src.utils.xlsx_utils import read_excel_data, count_filled_rows
from src.utils.docCheck_utils import update_document
from src.utils.dataEntry_utils import select_dropdown_by_value
from src.utils.assign_utils import assign_data
from src.utils.ui_utils import setup_webdriver, save_screenshot
from src.utils.autoScoring_utils import automate_scoring
from src.config import config

import time
import re
import sys
import random
import os
from faker import Faker

from src.utils.logging_utils import setup_logging
import logging

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser", default=config.get("browser", "edge"), help="Browser type")
    parser.add_argument("--log-level", default=config.get("log_level", "info"), help="Log level")
    parser.add_argument("--fast", action="store_true", help="Run in fast/headless mode")
    args, unknown = parser.parse_known_args()

    # Setup logging
    setup_logging(args.log_level)
    logger = logging.getLogger("LOSAuto.SupplementLOS")

    # Fast mode: reduced delays
    fast = args.fast
    def wait(seconds):
        """Adaptive sleep: reduced in fast mode."""
        time.sleep(max(0.5, seconds * 0.3) if fast else seconds)

    driver = setup_webdriver(args.browser, fast_mode=fast)

    # Setup path data from config
    xlsxdataPath = config.get("data_file", "Data/LOSData.xlsx")
    sheetEntryData = "Entry Data"
    sheetDocCheck = "Document Checking"
    sheetJobInfo = "Job Info"
    sheetAssignment = "Assignment"
    sheetEmergency = "Emergency Contact"
    sheetOtherOptions = "Other Options"
    sheetCard = "Card"
    sheetSupplement = "Supplement"

    # Retrieve Data
    fake = Faker('en_NZ')
    qrCodeData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 1)
    branchData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 2)
    channelData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 3)
    kodeProgramData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 4)
    jenisNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 5)
    salesIDData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 6)
    namaNasabahDataOri = read_excel_data(xlsxdataPath, sheetEntryData, 2, 7)
    namaNasabahData = namaNasabahDataOri + str(" ") + fake.first_name().upper()
    namaCetakArr = namaNasabahData.split(' ')
    namaCetakData = namaCetakArr[2] if len(namaCetakArr) > 2 else namaCetakArr[0]
    namaPanggilanData = namaCetakData
    jenisKelaminData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 10)
    tempatLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 12)  
    tanggalLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 13)
    bulanLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 14)
    tahunLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 15)
    eKTPData = str(random.randint(10**15, 10**16 - 1))
    statusPernikahanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 17)
    jumlahTanggunganData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 18)
    pendidikanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 19)
    namaIbuData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 20)
    identitasPajakData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 21)
    alamatRumahData1 = read_excel_data(xlsxdataPath, sheetEntryData, 2, 22)
    alamatRumahData2 = read_excel_data(xlsxdataPath, sheetEntryData, 2, 23)
    alamatRumahData3 = read_excel_data(xlsxdataPath, sheetEntryData, 2, 24)
    alamatRT = read_excel_data(xlsxdataPath, sheetEntryData, 2, 25)
    alamatRW = read_excel_data(xlsxdataPath, sheetEntryData, 2, 26)
    alamatLurah = read_excel_data(xlsxdataPath, sheetEntryData, 2, 27) 
    alamatCamat = read_excel_data(xlsxdataPath, sheetEntryData, 2, 28) 
    kotaData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 29)
    kodePosData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 30)
    telpRumahPrefixData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 31)
    telpRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 32)
    telpHPData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 33)
    statusRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 34)
    lamaMenempatiThnData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 35)
    lamaMenempatiBlnData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 36)
    NPWPData = str(random.randint(10**14, 10**15 - 1))
    emailData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 38)
    
    workbook = load_workbook(xlsxdataPath)
    sheet1 = workbook['Entry Data']
    sheet1.cell(row=2, column=8).value = namaPanggilanData
    sheet1.cell(row=2, column=9).value = namaCetakData
    sheet1.cell(row=2, column=16).value = eKTPData
    sheet1.cell(row=2, column=37).value = NPWPData
    workbook.save(xlsxdataPath)

    docCheckData1 = read_excel_data(xlsxdataPath, sheetDocCheck, 2, 1)
    docCheckData2 = read_excel_data(xlsxdataPath, sheetDocCheck, 3, 1)
    docCheckData3 = read_excel_data(xlsxdataPath, sheetDocCheck, 4, 1)

    namaPerusahaanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 1)
    kategoriPekerjaanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 2)
    bidangUsahaData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 3)
    lamaBekerjaThnData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 4)
    lamaBekerjaBlnData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 5)
    pangkatData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 6)
    pendapatanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 7)
    alamatKantor1 = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 8)
    alamatKantor2 = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 9)
    alamatKantor3 = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 10)
    alamatKantorRT = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 11)
    alamatKantorRW = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 12)
    alamatKantorLurah = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 13)
    alamatKantorCamat = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 14)
    alamatKantorPos = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 15)
    kotaKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 16)
    telpKantorPrefixData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 17)
    telpKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 18)

    namaEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 1)
    hubunganEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 2)
    alamatEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 3)
    kotaEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 4)
    posEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 5)
    telpRumahPrefixEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 6)
    telpRumahEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 7)

    tagihanData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 1)
    kartuUtamaData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 2)
    kartuSupplementData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 5)
    pilihanBSData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 3)
    noRekData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 4)

    networkData = read_excel_data(xlsxdataPath, sheetCard, 2, 1)
    productData = read_excel_data(xlsxdataPath, sheetCard, 2, 2)
    cardTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 3)
    plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 4)
    limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)
    potData = read_excel_data(xlsxdataPath, sheetCard, 2, 9)

    sendToData = read_excel_data(xlsxdataPath, sheetAssignment, 2, 1)
    sendToAppReviewer = read_excel_data(xlsxdataPath, sheetAssignment, 2, 2)

    try:
        start_time = time.time()
        driver.get("http://172.24.141.61/bricc/unittest.aspx")
        logger.info("Login")
        credentials = config.get_credentials()
        perform_login(driver, credentials.get("username", "cc_dam"))

        wait(2)
        logger.info("Generate QR Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=BAR|G&passurl&mntitle=Generate%20Barcode&row=8&barfontsize=40&barsize=100&imgwidth=250&borderwidth=0&altalign=1&padtop=20&rowheight=73&colwidth=350&cellspacing=0")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']").send_keys(branchData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_btn_gen']").click()
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        save_screenshot(driver, '2.png')
        image_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='mainPanel_tbl']/tbody/tr[2]/td[1]/img")))
        image_url = image_element.get_attribute('src')
        qrCode = image_url.split("data=")[1].split("&")[0]
        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=1).value = qrCode
        workbook.save(xlsxdataPath)
        driver.switch_to.default_content()
        logger.info("Generate QR Step End")

        logger.info("Initial Data Entry Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE&passurl&mntitle=Initial%20Data%20Entry&tc=1.0&stg=DE")
        WebDriverWait(driver, 620).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 620).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_EKTP_NO']"))).send_keys(eKTPData)
        WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_df_BARCODE']"))).send_keys(qrCodeData)
        WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']"))).send_keys(branchData)
        WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        wait(2)
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_df_SALES2']"))).click()
        wait(2)
        WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_df_SALES2']"))).send_keys("42352")
        wait(2)
        selectChannel = Select(driver.find_element(By.ID, "mainPanel_channelPanel_df_CH_CODE"))
        selectChannel.select_by_value(str(channelData))
        wait(2)
        WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, "mainPanel_programPanel_df_PR_CODE")))
        selectProgram = Select(driver.find_element(By.ID, "mainPanel_programPanel_df_PR_CODE"))
        selectProgram.select_by_value(str(kodeProgramData))
        wait(3)
        WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, "mainPanel_custcatPanel_df_CUSTCATID")))
        selectJenisNasabah = Select(driver.find_element(By.ID, "mainPanel_custcatPanel_df_CUSTCATID"))
        selectJenisNasabah.select_by_value(str(jenisNasabahData))
        wait(3)
        WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NMFIRST']").send_keys(namaNasabahData)
        wait(3)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNDATE_DD']").send_keys(tanggalLahirData)
        selectBulanLahir = Select(driver.find_element(By.ID, "mainPanel_CU_BORNDATE_MM"))
        selectBulanLahir.select_by_value(str(bulanLahirData))
        wait(3)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNDATE_YY']").send_keys(tahunLahirData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_MMNMFIRST']").send_keys(namaIbuData)
        save_screenshot(driver, '3.png')
        driver.find_element(By.ID,"mainPanel_btn_save").click()
        WebDriverWait(driver, 30).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert_text = alert.text
        pattern = r'\bKNP\w+'
        match = re.search(pattern, alert_text)
        noAplikasiData = match.group(0) if match else "UNKNOWN"
        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=39).value = noAplikasiData
        workbook.save(xlsxdataPath)
        alert.accept()
        driver.switch_to.default_content()
        wait(3)
        logger.info("Initial Data Entry Step End")

        logger.info("Document Checking Step Start")
        update_appflag(noAplikasiData, ap_currtrcode="1.1", ap_lasttrcode="1.0")
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE|D&passurl&mntitle=Document%20Checking&tc=1.1&regno={noAplikasiData}&stg=DE")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        docCheckDataList = [docCheckData1, docCheckData2, docCheckData3]
        for docCheckData in docCheckDataList:
            WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "griddoc_header7_Button4"))).click()
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            update_document(driver, docCheckData)
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            wait(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        wait(3)
        save_screenshot(driver, '4.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "updPanel_btnUpdate"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 10).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
            except:
                pass
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            wait(2)
        driver.switch_to.default_content()
        logger.info("Document Checking Step End")

        logger.info("Data Entry Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV|DE&passurl&mntitle=Data%20Entry%20Assignment&li1=LR|A|SPV|DE|00&li2=LR|A|SPV|DE|01&tc1=3.0&tc2=3.1&atype=DE&stg=DE&xlcode=DE")
        WebDriverWait(driver, 320).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        save_screenshot(driver, '5.png')
        assign_data(driver, "mainPanel_sendtoPanel_SendTo", noAplikasiData, sendToData)
        logger.info("Data Entry Assignment Step End")

        logger.info("Data Entry Step Start")
        update_user(noAplikasiData)
        wait(3)
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={noAplikasiData}&stg=DE")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Personal Info']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NICKNM']").send_keys(namaPanggilanData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NMCARD']").send_keys(namaCetakData)
        driver.find_element(By.XPATH, f"(//input[@id='mainPanel_CU_GENDER_CODE_{str(jenisKelaminData)}'])[1]").click()
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNPLACE']").send_keys(tempatLahirData)
        driver.find_element(By.ID, "mainPanel_CU_EMAIL").send_keys(emailData)
        Select(driver.find_element(By.ID, "mainPanel_CU_MR_CODE")).select_by_value(str(statusPernikahanData))
        driver.find_element(By.ID, "mainPanel_CU_NPWP").send_keys(NPWPData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_CHILD']").send_keys(jumlahTanggunganData)
        Select(driver.find_element(By.ID, "mainPanel_CU_ED_CODE")).select_by_value(str(pendidikanData))
        Select(driver.find_element(By.ID, "mainPanel_CU_TAXIDFLAG")).select_by_value(str(identitasPajakData))
        driver.find_element(By.ID, "mainPanel_CU_HMADDR1").send_keys(alamatRumahData1)
        driver.find_element(By.ID, "mainPanel_CU_HMADDR2").send_keys(alamatRumahData2)
        driver.find_element(By.ID, "mainPanel_CU_HMADDR3").send_keys(alamatRumahData3)
        driver.find_element(By.ID, "mainPanel_CU_HMRT").send_keys(alamatRT)
        driver.find_element(By.ID, "mainPanel_CU_HMRW").send_keys(alamatRW)
        driver.find_element(By.ID, "mainPanel_CU_HMLURAH").send_keys(alamatLurah)
        driver.find_element(By.ID, "mainPanel_CU_HMCAMAT").send_keys(alamatCamat)
        driver.find_element(By.ID, "mainPanel_CU_HMCITY").send_keys(kotaData)
        driver.find_element(By.ID, "mainPanel_CU_HMZIPCODE_txtID").send_keys(kodePosData)
        driver.find_element(By.ID, "mainPanel_CU_HMPHONEAREA").send_keys(telpRumahPrefixData)
        driver.find_element(By.ID, "mainPanel_CU_HMPHONE").send_keys(telpRumahData)
        driver.find_element(By.ID, "mainPanel_CU_HPNO").send_keys(telpHPData)
        Select(driver.find_element(By.ID, "mainPanel_CU_HMSTA_CODE")).select_by_value(str(statusRumahData))
        driver.find_element(By.ID, "mainPanel_CU_HMLIVESINCEYY").send_keys(lamaMenempatiThnData)
        driver.find_element(By.ID, "mainPanel_CU_HMLIVESINCEMM").send_keys(lamaMenempatiBlnData)
        save_screenshot(driver, '6.png')
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        WebDriverWait(driver, 10).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        wait(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        driver.switch_to.default_content()

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Job Info']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        driver.find_element(By.ID, "mainPanel_CU_COMPNAME").send_keys(namaPerusahaanData)
        Select(driver.find_element(By.ID, "mainPanel_CU_JOB_CODE")).select_by_value(str(kategoriPekerjaanData))
        driver.find_element(By.ID, "mainPanel_CU_BUSSINESSDESC").send_keys(bidangUsahaData)
        driver.find_element(By.ID, "mainPanel_CU_WORKSINCEYY").send_keys(lamaBekerjaThnData)
        driver.find_element(By.ID, "mainPanel_CU_WORKSINCEMM").send_keys(lamaBekerjaBlnData)
        driver.find_element(By.ID, "mainPanel_CU_OFTITLE").send_keys(pangkatData)
        driver.find_element(By.ID, "mainPanel_CU_INCOME").send_keys(pendapatanData)
        driver.find_element(By.ID, "mainPanel_CU_OFADDR1").send_keys(alamatKantor1)
        driver.find_element(By.ID, "mainPanel_CU_OFADDR2").send_keys(alamatKantor2)
        driver.find_element(By.ID, "mainPanel_CU_OFADDR3").send_keys(alamatKantor3)
        driver.find_element(By.ID, "mainPanel_CU_OFRT").send_keys(alamatKantorRT)
        driver.find_element(By.ID, "mainPanel_CU_OFRW").send_keys(alamatKantorRW)
        driver.find_element(By.ID, "mainPanel_CU_OFLURAH").send_keys(alamatKantorLurah)
        driver.find_element(By.ID, "mainPanel_CU_OFCAMAT").send_keys(alamatKantorCamat)
        driver.find_element(By.ID, "mainPanel_CU_OFCITY").send_keys(kotaKantorData)
        driver.find_element(By.ID, "mainPanel_CU_OFZIPCODE_txtID").send_keys(alamatKantorPos)
        driver.find_element(By.ID, "mainPanel_CU_OFPHONEAREA").send_keys(telpKantorPrefixData)
        driver.find_element(By.ID, "mainPanel_CU_OFPHONE").send_keys(telpKantorData)
        save_screenshot(driver, '7.png')
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        WebDriverWait(driver, 999).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        driver.switch_to.default_content()

        # ========== SUPPLEMENT DATA ENTRY ==========
        logger.info("Supplement Data Entry Start")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Supplement']"))).click()
        
        # Count filled rows in Supplement sheet
        workbook = load_workbook(xlsxdataPath)
        sheet = workbook['Supplement']
        filled_rows_supplement = count_filled_rows(sheet)
        workbook.close()

        # Access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        wait(1)
        
        # Loop through each supplement row
        for row in range(2, filled_rows_supplement + 1):
            # Generate random name for supplement card holder
            namaSupplement = fake.name_male()
            
            # Read supplement data from Excel
            tanggalLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, row, 2)
            bulanLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, row, 3)
            tahunLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, row, 4)
            noHPSupplement = read_excel_data(xlsxdataPath, sheetSupplement, row, 8)

            # Save generated name to Excel
            workbook = load_workbook(xlsxdataPath)
            sheet_supp = workbook['Supplement']
            sheet_supp.cell(row=row, column=1).value = namaSupplement
            workbook.save(xlsxdataPath)
            workbook.close()

            # Click Add Supplement button
            WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "GridViewSuppInfo_header6_Button5"))).click()
            wait(3)
            
            # Fill supplement form
            driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_NAME").send_keys(namaSupplement)
            driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_DD").send_keys(tanggalLahirSupplement)
            
            WebDriverWait(driver, 180).until(EC.presence_of_element_located((By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_MM")))
            Select(driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_MM")).select_by_value(str(bulanLahirSupplement))
            wait(1)
            
            driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_YY").send_keys(tahunLahirSupplement)
            driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_HPNO").send_keys(noHPSupplement)
            
            save_screenshot(driver, '8.png')

            # Save supplement
            WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "PopupSuppInfo_PNL_SuppInfo_BTN_SAVE"))).click()
            wait(3)
            logger.info(f"Supplement {row - 1} added: {namaSupplement}")

        save_screenshot(driver, '9.png')
        driver.switch_to.default_content()
        logger.info("Supplement Data Entry End")

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Emergency Contact']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.ID, "mainPanel_CU_EMNMFIRST")))
        driver.find_element(By.ID, "mainPanel_CU_EMNMFIRST").send_keys(namaEmergency)
        selectRel = Select(driver.find_element(By.ID, "mainPanel_CU_EMRE_CODE"))
        selectRel.select_by_value(str(hubunganEmergency))
        driver.find_element(By.ID, "mainPanel_CU_EMHMADDR1").send_keys(alamatEmergency)
        driver.find_element(By.ID, "mainPanel_CU_EMHMCITY").send_keys(kotaEmergency)
        driver.find_element(By.ID, "mainPanel_CU_EMHMZIPCODE_txtID").send_keys(posEmergency)
        driver.find_element(By.ID, "mainPanel_CU_EMHMPHONEAREA").send_keys(telpRumahPrefixEmergency)
        driver.find_element(By.ID, "mainPanel_CU_EMHMPHONE").send_keys(telpRumahEmergency)
        save_screenshot(driver, '10.png')
        WebDriverWait(driver, 9999).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        WebDriverWait(driver, 9999).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        wait(3)
        driver.switch_to.default_content()

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Other Options']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        driver.find_element(By.ID, f"mainPanel_BILLING_DLV_TYPE_{tagihanData}").click()
        driver.find_element(By.ID, f"mainPanel_MAINCARD_DLV_TYPE_{kartuUtamaData}").click()
        driver.find_element(By.ID, f"mainPanel_SUPPCARD_DLV_TYPE_{kartuSupplementData}").click()
        Select(driver.find_element(By.ID, "mainPanel_BILLING_STAT_TYPE")).select_by_value(str(pilihanBSData))
        driver.find_element(By.ID, "mainPanel_AD_ACCNUM").send_keys(noRekData)
        save_screenshot(driver, '11.png')
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        wait(3)
        WebDriverWait(driver, 50).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        wait(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        driver.switch_to.default_content()

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Main']"))).click()
        wait(3)

        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "mainPanel_GV_PRODUCT_header5_BTN_NEW"))).click()
        wait(3)
        Select(driver.find_element(By.XPATH, "//*[@id='popupDetail_panelDetail_df_NETWORKID']")).select_by_value(networkData)
        wait(3)
        Select(driver.find_element(By.ID, "popupDetail_panelDetail_df_PRODUCTID")).select_by_value(productData)
        wait(3)
        if (networkData in ["G", "J", "P"] and productData == "P") or \
           (networkData == "M" and productData in ["C", "G", "P"]) or \
           (networkData == "V"):
            select_dropdown_by_value(driver, "popupDetail_panelDetail_cardtypePanel_df_CARDTYPEID", cardTypeData)
        wait(3)
        Select(driver.find_element(By.ID, "popupDetail_panelDetail_plastictypePanel_df_PLASTICID")).select_by_value(str(plasticTypeData))
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "popupDetail_panelDetail_Button3"))).click()
        wait(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.ID, 'popupDetail_CSD-1')))
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={noAplikasiData}&stg=DE")
        
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '12.png')
        WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        WebDriverWait(driver, 620).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        wait(3)
        WebDriverWait(driver, 620).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        driver.switch_to.default_content()
        logger.info("Data Entry Step End")

        logger.info("DE Validation Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DE%20Validation%20Assignment&li1=LR|SPV|DE|00&li2=LR|SPV|01&tc1=3.5.0&tc2=3.5.1&stg=VDE")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", noAplikasiData, sendToData)
        logger.info("DE Validation Assignment Step End")

        logger.info("Data Entry Validation Step Start")
        update_user(noAplikasiData)
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VDE&passurl&mntitle=Data%20Entry%20Validation&tc=3.5.1&regno={noAplikasiData}&stg=DE")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '13.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Data Entry Validation Step End")

        """
        Scoring
        """
        host = '172.24.141.33'
        usernameScoring = 'administrator'
        passwordScoring = 'c@rdl1nk'
        pathWinSCP = r'C:/Program Files (x86)/WinSCP/WinSCP.exe'

        logger.info("Proses Scoring Mulai")
        appFlagAfter = automate_scoring(pathWinSCP, noAplikasiData, host, usernameScoring, passwordScoring)

        if appFlagAfter in ["9.2.1", "9.2.3"]:
            sys.exit(f"Program dihentikan karena appFlagAfter {appFlagAfter}")
        logger.info("Proses Scoring Selesai")

        logger.info("Credit Reviewer Assignment Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=Credit%20Reviewer%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=5.7.0&tc2=5.7.1&stg=ANLMINI")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", noAplikasiData, sendToAppReviewer)
        logger.info("Credit Reviewer Assignment End")

        logger.info("Credit Reviewer Start")
        update_user(noAplikasiData)
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=ANLMINI&passurl&mntitle=Credit%20Reviewer&tc=5.7.1&regno={noAplikasiData}&stg=ANL")
        wait(10)
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Decision']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "mainPanel_GV_PRODUCT_cell0_6_BTN_EDIT"))).click()
        wait(3)
        driver.find_element(By.ID, "popupDetail_panelDetail_df_CP_APRLIMIT").send_keys(str(limitData))
        select_dropdown_by_value(driver, "popupDetail_panelDetail_potPanel_df_POT_CODE", potData)
        save_screenshot(driver, '14.png')
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "popupDetail_panelDetail_Button6"))).click()
        wait(3)
        WebDriverWait(driver, 620).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        
        # ========== SUPPLEMENT CARD APPROVAL IN CREDIT REVIEWER ==========
        logger.info("Processing Supplement Card Approvals...")
        
        for row in range(2, filled_rows_supplement + 1):
            # Read supplement approval data
            limitSupplementData = read_excel_data(xlsxdataPath, sheetSupplement, row, 7)
            kartuBasicData = read_excel_data(xlsxdataPath, sheetSupplement, row, 6)
            namaLengkapSupplementData = read_excel_data(xlsxdataPath, sheetSupplement, row, 1)
            
            # Click edit button for supplement row
            cell_index = row - 2
            wait(3)
            WebDriverWait(driver, 180).until(
                EC.element_to_be_clickable((By.ID, f"mainPanel_GridViewSuppInfo_cell{cell_index}_7_Button1"))
            ).click()
            
            # Wait for popup to appear
            WebDriverWait(driver, 620).until(EC.visibility_of_element_located((By.ID, 'PopupSuppInfo_CSD-1')))
            WebDriverWait(driver, 180).until(
                EC.presence_of_element_located((By.ID, 'PopupSuppInfo_PNL_SuppInfo_mainBasicPanel_SUP_MAINCARD'))
            )
            wait(3)
            
            # Select main card for supplement
            Select(driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_mainBasicPanel_SUP_MAINCARD")).select_by_index(int(kartuBasicData))
            wait(1)
            
            # Fill approval limit
            driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_APRLIMIT").send_keys(str(limitSupplementData))
            wait(1)
            
            # Fill card name
            driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_NMCARD").send_keys(str(namaLengkapSupplementData))
            wait(1)
            
            # Select plastic type
            plastic_index = int(plasticTypeData) - 1
            WebDriverWait(driver, 180).until(
                EC.element_to_be_clickable((By.ID, f"PopupSuppInfo_PNL_SuppInfo_SUP_PLASTICID_{plastic_index}"))
            ).click()
            wait(1)
            
            Select(driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_suppotPanel_SUP_POT_CODE")).select_by_value(str(potData))
            wait(1)
            
            save_screenshot(driver, '15.png')

            # Save supplement approval
            WebDriverWait(driver, 180).until(
                EC.element_to_be_clickable((By.ID, "PopupSuppInfo_PNL_SuppInfo_Button2"))
            ).click()
            wait(5)
            
            logger.info(f"Supplement {row - 1} approved: {namaLengkapSupplementData}")

        Select(driver.find_element(By.ID, 'mainPanel_decPanel_dfa_AP_NEXTTRBY')).select_by_value("DIRTONI")
        save_screenshot(driver, '16.png')
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_decPanel_btn_update']"))).click()
        WebDriverWait(driver, 999).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        driver.switch_to.default_content()
        logger.info("Credit Reviewer Step End")
        
        wait(5)

        logger.info("Approval Step Start")
        update_user(noAplikasiData)
        wait(3)
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=APRV&passurl&mntitle=Approval&tc=7.1&regno={noAplikasiData}&stg=APRV")
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Decision']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '17.png')
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.ID, "mainPanel_decPanel_btn_appr"))).click()
        WebDriverWait(driver, 999).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
        driver.switch_to.default_content()
        logger.info("Approval Step End")

        logger.info("Proses Selesai")
        logger.info(f"Time: {int(time.time() - start_time)}s")

    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()

if __name__ == "__main__":
    main()