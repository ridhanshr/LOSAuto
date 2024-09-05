from openpyxl import load_workbook

def read_excel_data(file_path, sheet_name, row, column):
    """
    Perform Read Excel Data.

    Parameters:
    - file_path: Path xlsx file.
    - sheet_name: Sheet name.
    - row: row
    - column: column

    Returns:
    - Read data from excel spesific row and column
    """
    # Load the Excel workbook
    workbook = load_workbook(file_path)
    # Select the sheet by name
    sheet = workbook[sheet_name]
    # Retrieve data from specified row and column
    cell_value = sheet.cell(row=row, column=column).value
    return cell_value

def count_filled_rows(sheet):
    max_row = sheet.max_row
    max_col = sheet.max_column
    filled_rows = 0
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if sheet.cell(row=row, column=col).value not in (None, ''):
                filled_rows += 1
                break 
    
    return filled_rows


# # Membuka file Excel
# workbook = load_workbook('Data/LOSData.xlsx')
# sheet = workbook['Card']
# # Contoh penggunaan
# filled_rows = count_filled_rows(sheet)
# print(f"Jumlah baris yang diisi: {filled_rows}")