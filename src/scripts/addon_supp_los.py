from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from src.utils.db_utils import update_appflag, update_user, update_accnum, check_appflag, update_applicant, update_norek, update_paramater_key
from src.utils.login_utils import perform_login
from src.utils.xlsx_utils import read_excel_data, count_filled_rows
from src.utils.docCheck_utils import update_document
from src.utils.dataEntry_utils import select_dropdown_by_value
from src.utils.assign_utils import assign_data
from src.utils.ui_utils import setup_webdriver, save_screenshot
from src.config import config

import time
import datetime
import re
import sys
import random
import os
from faker import Faker

from src.utils.logging_utils import setup_logging
import logging

def main():
    # Parse command line arguments
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser", default=config.get("browser", "edge"), help="Browser type")
    parser.add_argument("--log-level", default=config.get("log_level", "info"), help="Log level")
    args, unknown = parser.parse_known_args()

    # Setup logging
    setup_logging(args.log_level)
    logger = logging.getLogger("LOSAuto.AddonSupp")

    # Setup webdriver based on argument or config
    driver = setup_webdriver(args.browser)

    # Setup path data from config
    xlsxdataPath = config.get("data_file", "Data/LOSData.xlsx")
    sheetEntryData = "Entry Data"
    sheetEntryDataAddOn = "Entry Data Add On"
    sheetSupplement = "Supplement"
    sheetDocCheck = "Document Checking"
    sheetAssignment = "Assignment"
    sheetCard = "Card"

    # Retrieve Data
    # Data Entry Data
    fake = Faker('en_NZ')
    qrCodeData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 1)
    branchData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 2)
    channelData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 3)
    kodeProgramData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 4)
    jenisNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 5)

    # Entry Data Add On
    noKartu = read_excel_data(xlsxdataPath, sheetEntryDataAddOn, 2, 1)

    # Document Check Data
    docCheckData1 = read_excel_data(xlsxdataPath, sheetDocCheck, 2, 1)
    docCheckData3 = read_excel_data(xlsxdataPath, sheetDocCheck, 4, 1)

    #Card Data
    plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 4)
    limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)
    namaNasabahAddOn = fake.name_male()
    potData = read_excel_data(xlsxdataPath, sheetCard, 2, 9)
    
    workbook = load_workbook(xlsxdataPath)
    sheetCardObj = workbook['Card']
    sheetCardObj.cell(row=2, column=6).value = namaNasabahAddOn
    workbook.save(xlsxdataPath)

    # Assignment Data
    sendToData = read_excel_data(xlsxdataPath, sheetAssignment, 2, 1)

    try:
        start_time = time.time()
        """
        login
        """
        driver.get("http://172.24.141.61/bricc/unittest.aspx")
        logger.info("Login")
        credentials = config.get_credentials()
        perform_login(driver, credentials.get("username", "cc_dam"))

        """
        generate QR
        """
        time.sleep(10)
        logger.info("Generate QR Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=BAR|G&passurl&mntitle=Generate%20Barcode&row=8&barfontsize=40&barsize=100&imgwidth=250&borderwidth=0&altalign=1&padtop=20&rowheight=73&colwidth=350&cellspacing=0")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']").send_keys(branchData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_btn_gen']").click()
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        image_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='mainPanel_tbl']/tbody/tr[2]/td[1]/img")))
        image_url = image_element.get_attribute('src')
        qrCode = image_url.split("data=")[1].split("&")[0]
        qrCodeSub = re.sub(r'^(.....)1(.*)$', r'\g<1>0\g<2>', qrCode)
        workbook = load_workbook(xlsxdataPath)
        sheet1Idx = workbook['Entry Data']
        sheet1Idx.cell(row=2, column=1).value = qrCodeSub
        save_screenshot(driver, '2.png')
        workbook.save(xlsxdataPath)
        driver.switch_to.default_content()
        logger.info("Generate QR Step End")

        """
        initial data entry add on
        """
        logger.info("initial data add on")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE|A&passurl&mntitle=Add-On%20Supplement&tc=1.0&stg=DE")
        time.sleep(1)
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "(//input[@name='mainPanel$ctl04'])[1]"))).click()
        time.sleep(1)
        driver.find_element(By.XPATH, "//input[@id='PopFindBasic_PNL_FindBasic_fCORE_ID']").send_keys(noKartu)
        time.sleep(2)
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Find']"))).click()
        save_screenshot(driver, '3.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "PopFindBasic_PNL_FindBasic_GridViewX_cell0_8_Button4"))).click()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        driver.find_element(By.CSS_SELECTOR, "#mainPanel_branchPanel_df_BRANCHID_txtID").clear()
        time.sleep(5)
        driver.find_element(By.CSS_SELECTOR, "#mainPanel_branchPanel_df_BRANCHID_txtID").send_keys(branchData)
        time.sleep(5)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_df_BARCODE']").send_keys(qrCodeSub)
        time.sleep(5)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        Select(driver.find_element(By.ID, "mainPanel_channelPanel_df_CH_CODE")).select_by_value(str(channelData))
        time.sleep(5)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        Select(driver.find_element(By.ID, "mainPanel_programPanel_df_PR_CODE")).select_by_value(str(kodeProgramData))
        time.sleep(5)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        Select(driver.find_element(By.ID, "mainPanel_custcatPanel_df_CUSTCATID")).select_by_value(str(jenisNasabahData))
        time.sleep(5)
        save_screenshot(driver, '4.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_Button2']"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(5)
            except:
                pass
        driver.switch_to.default_content()

        """
        Document Checking
        """
        print("Document Checking Step Start")
        time.sleep(2)
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        valueNoAplikasi = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "UC_GeneralInfo1_l1"))).text
        update_paramater_key(valueNoAplikasi, str(channelData), str(jenisNasabahData), str(kodeProgramData))
        time.sleep(2)
        workbook = load_workbook(xlsxdataPath)
        sheet1App = workbook['Entry Data']
        sheet1App.cell(row=2, column=39).value = valueNoAplikasi
        workbook.save(xlsxdataPath)
        docCheckDataList = [docCheckData1, docCheckData3]
        for docCheckData in docCheckDataList:
            WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "griddoc_header7_Button4"))).click()
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            update_document(driver, docCheckData)
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(3)
        save_screenshot(driver, '5.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "updPanel_btnUpdate"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 30).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(2)
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Document Checking Step End")

        """
        Data Entry Assignment
        """
        logger.info("Data Entry Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV|DE&passurl&mntitle=Data%20Entry%20Assignment&li1=LR|A|SPV|DE|00&li2=LR|A|SPV|DE|01&tc1=3.0&tc2=3.1&atype=DE&stg=DE&xlcode=DE")
        WebDriverWait(driver, 320).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_sendtoPanel_SendTo", valueNoAplikasi, sendToData)
        logger.info("Data Entry Assignment Step End")

        """
        Data Entry
        """
        logger.info("Data Entry Step Start")
        update_user(valueNoAplikasi)
        time.sleep(3)
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={valueNoAplikasi}&stg=DE")
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Supplement']"))).click()
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        time.sleep(1)
        namaSupplement = fake.name_male() + "BRIMO"
        bulanLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 3)
        tahunLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 4)
        noHPSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 8)
        limitSupplementData = read_excel_data(xlsxdataPath, sheetSupplement, 2, 7)
        workbook = load_workbook(xlsxdataPath)
        sheetSupp = workbook['Supplement']
        sheetSupp.cell(row=1, column=2).value = namaSupplement
        workbook.save(xlsxdataPath)
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "GridViewSuppInfo_header6_Button5"))).click()
        time.sleep(3)
        driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_NAME").send_keys(namaSupplement)
        driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_DD").send_keys(bulanLahirSupplement)
        Select(driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_MM")).select_by_value(str(bulanLahirSupplement))
        time.sleep(3)
        driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_YY").send_keys(tahunLahirSupplement)
        driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_LIMIT").send_keys(limitSupplementData)
        driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_HPNO").send_keys(noHPSupplement)
        driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_NMCARD").send_keys(namaSupplement)
        save_screenshot(driver, '6.png')
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "PopupSuppInfo_PNL_SuppInfo_BTN_SAVE"))).click()
        time.sleep(3)
        driver.switch_to.default_content()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Main']"))).click()
        update_accnum(valueNoAplikasi)
        time.sleep(1)
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '7.png')
        WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 620).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(3)
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Data Entry Step End")

        """
        DE Validation Assignment
        """
        logger.info("DE Validation Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DE%20Validation%20Assignment&li1=LR|SPV|DE|00&li2=LR|SPV|01&tc1=3.5.0&tc2=3.5.1&stg=VDE")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, sendToData)
        logger.info("DE Validation Assignment Step End")

        """
        DE Validation
        """
        time.sleep(3)
        print("Data Entry Validation Step Start")
        update_user(valueNoAplikasi)
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VDE&passurl&mntitle=Data%20Entry%20Validation&tc=3.5.1&regno={valueNoAplikasi}&stg=DE")
        time.sleep(3)
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '8.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(5)
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Data Entry Validation Step End")

        """
        Verphone Add Basic Supplement Assignment
        """
        update_appflag(valueNoAplikasi, '5.5.0', '3.5.1')
        time.sleep(3)
        logger.info("Verphone Add Basic Supplement Assignment Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=VerPhone%20Supplement%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=5.5.0&tc2=5.5.1&stg=VERSUP")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, 'MY_BRI_APPROVAL')
        logger.info("Verphone Add Basic Supplement Assignment End")

        """
        Verphone Supplement
        """
        update_user(valueNoAplikasi)
        logger.info("Verphone Supplement Start")
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VERSUP&passurl&mntitle=VerPhone%20Supplement&tc=5.5.1&regno={valueNoAplikasi}")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '9.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_Button6"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(5)
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Verphone Supplement End")

        """
        Data Maintenance Assignment
        """
        logger.info("Data Maintenance Assignment Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=Data%20Maintenance%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=6.5.0&tc2=6.5.1&stg=DMSUP")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, 'MY_BRI_APPROVAL')
        logger.info("Data Maintenance Assignment End")

        """
        Data Maintenance for Suplement
        """
        update_user(valueNoAplikasi)
        logger.info("Data Maintenance for Suplement Start")
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DMSUP&passurl&mntitle=Data%20Maintenance%20for%20Supplement&tc=6.5.1&regno={valueNoAplikasi}")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        time.sleep(1)
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "GridViewSuppInfo_cell0_6_Button1"))).click()
        time.sleep(3)
        plastic_index = int(plasticTypeData) - 1
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, f"PopupSuppInfo_PNL_SuppInfo_SUP_PLASTICID_{str(plastic_index)}"))).click()
        Select(driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_POT_CODE")).select_by_value(str(potData))
        dateTimeNow = datetime.datetime.now()
        driver.find_element(By.ID, 'PopupSuppInfo_PNL_SuppInfo_SUP_POT_EXPMM').send_keys("0" + str(dateTimeNow.month))
        driver.find_element(By.ID, 'PopupSuppInfo_PNL_SuppInfo_SUP_POT_EXPYY').send_keys(str(dateTimeNow.year + 5))
        time.sleep(2)
        save_screenshot(driver, '10.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "PopupSuppInfo_PNL_SuppInfo_Button1"))).click()
        save_screenshot(driver, '11.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_btn_appr"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(5)
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Data Maintenance for Suplement End")

        """
        Data Maintenance for Supplement Approval
        """
        logger.info("Data Maintenance for Supplement Approval Start")
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DMSUP&passurl&mntitle=Data%20Maintenance%20for%20Supplement%20Approval&tc=6.5.2&regno={valueNoAplikasi}")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '12.png')
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_btn_appr"))).click()
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(5)
            except:
                pass
        driver.switch_to.default_content()
        logger.info("Data Maintenance for Supplement Approval End")

        logger.info("Proses Onboarding Data Add On Basic LOS Selesai")
        logger.info(f"Time: {int(time.time() - start_time)}s")

    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()

if __name__ == "__main__":
    main()