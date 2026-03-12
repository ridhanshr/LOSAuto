from openpyxl import load_workbook
import os

file_path = 'Data/LOSData.xlsx'
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Insert a new column at index 1
        sheet.insert_cols(1)
        # Set the header for the new column
        sheet.cell(row=1, column=1).value = 'ID'
        
        # Populate IDs for existing data if any (assuming row 1 is header)
        for row in range(2, sheet.max_row + 1):
            if any(sheet.cell(row=row, column=col).value for col in range(2, sheet.max_column + 1)):
                if sheet_name == 'Entry Data':
                     # For Entry Data, give unique IDs 1, 2, 3...
                     sheet.cell(row=row, column=1).value = row - 1
                else:
                    # For other sheets, we'll leave it to the user or for now put 1
                    # as a placeholder if there is data
                     sheet.cell(row=row, column=1).value = 1
            
    wb.save(file_path)
    print("Excel modification complete.")
else:
    print("File not found.")
