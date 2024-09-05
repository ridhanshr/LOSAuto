from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from utils.db_utils import update_appflag, update_user, update_norek,check_appflag
from utils.login_utils import perform_login
from utils.xlsx_utils import read_excel_data
from utils.docCheck_utils import update_document
from utils.dataEntry_utils import select_dropdown_by_value
from utils.assign_utils import assign_data
from utils.autoScoring_utils import automate_scoring

import time
import re
import sys
import random

"""
setup webdriver
"""
# Chrome
chrome_path = r'webDriver/chromedriver.exe'
service = Service(executable_path=chrome_path)
chrome_options = Options()
chrome_options.add_argument("--start-maximized")
# chrome_options.add_argument("--incognito")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--disable-software-rasterizer")
driver = webdriver.Chrome(service=service, options=chrome_options)

# Edge
# service = Service(executable_path="webDriver/msedgedriver.exe")
# edge_options = Options()
# edge_options.add_argument("--InPrivate")
# driver = webdriver.Edge(service=service, options=edge_options)
# driver.maximize_window()

"""
setup data
"""
# Setup path data
xlsxdataPath = r'Data/LOSData.xlsx'
sheetEntryData = "Entry Data"
sheetDocCheck = "Document Checking"
sheetJobInfo = "Job Info"
sheetAssignment = "Assignment"
sheetEmergency = "Emergency Contact"
sheetOtherOptions = "Other Options"
sheetCard = "Card"

# Retrieve Data
# Data Entry Data
qrCodeData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 1)
branchData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 2)
channelData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 3)
kodeProgramData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 4)
jenisNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 5)
salesIDData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 6)
namaNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 7)
namaPanggilanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 8)
namaCetakData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 9)
jenisKelaminData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 10)
tempatLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 12)
tanggalLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 13)
bulanLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 14)
tahunLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 15)
eKTPData = str(random.randint(10**15, 10**16 - 1))
statusPernikahanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 17)
jumlahTanggunganData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 18)
pendidikanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 19)
namaIbuData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 20)
identitasPajakData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 21)
alamatRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 22)
kotaData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 23)
kodePosData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 24)
telpRumahPrefixData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 25)
telpRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 26)
telpHPData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 27)
statusRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 28)
lamaMenempatiThnData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 29)
lamaMenempatiBlnData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 30)
NPWPData = str(random.randint(10**14, 10**15 - 1))
emailData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 32)
noAplikasiData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 33)

workbook = load_workbook(xlsxdataPath)
sheet1 = workbook['Entry Data']
sheet1.cell(row=2, column=16).value = eKTPData
sheet1.cell(row=2, column=31).value = NPWPData
workbook.save(xlsxdataPath)

# Document Check Data
docCheckData1 = read_excel_data(xlsxdataPath, sheetDocCheck, 2, 1)
docCheckData2 = read_excel_data(xlsxdataPath, sheetDocCheck, 3, 1)
docCheckData3 = read_excel_data(xlsxdataPath, sheetDocCheck, 4, 1)

# Job Info Data
namaPerusahaanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 1)
kategoriPekerjaanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 2)
bidangUsahaData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 3)
lamaBekerjaThnData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 4)
lamaBekerjaBlnData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 5)
pangkatData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 6)
pendapatanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 7)
alamatKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 8)
kotaKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 9)
telpKantorPrefixData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 10)
telpKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 11)

# Emergency Contact Data
namaEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 1)
hubunganEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 2)
alamatEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 3)
kotaEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 4)
telpRumahPrefixEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 5)
telpRumahEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 6)

#Other Options Data
tagihanData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 1)
kartuUtamaData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 2)
pilihanBSData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 3)
noRekData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 4)

#Card Data
networkData = read_excel_data(xlsxdataPath, sheetCard, 2, 1)
productData = read_excel_data(xlsxdataPath, sheetCard, 2, 2)
cardTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 3)
plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 4)
limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)

# Assignment Data
sendToData = read_excel_data(xlsxdataPath, sheetAssignment, 2, 1)
sendToAppReviewer = read_excel_data(xlsxdataPath, sheetAssignment, 2, 2)

start_time = time.time()
"""
login
"""
# access los
driver.get("http://172.24.141.61/bricc/unittest.aspx")
print("Login")
perform_login(driver)

"""
generate QR
"""
print("Generate QR Step Start")
driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=BAR|G&passurl&mntitle=Generate%20Barcode&row=8&barfontsize=40&barsize=100&imgwidth=250&borderwidth=0&altalign=1&padtop=20&rowheight=73&colwidth=350&cellspacing=0")

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# access element iframe
driver.find_element(By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']").send_keys(branchData)
driver.find_element(By.XPATH, "//input[@id='mainPanel_btn_gen']").click()
WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

image_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='mainPanel_tbl']/tbody/tr[2]/td[1]/img")))
image_url = image_element.get_attribute('src')
qrCode = image_url.split("data=")[1].split("&")[0]
workbook = load_workbook(xlsxdataPath)
sheet1 = workbook['Entry Data']
sheet1.cell(row=2, column=1).value = qrCode
workbook.save(xlsxdataPath)

# exit from iframe 
driver.switch_to.default_content()

print("Generate QR Step End")

"""
Initial Data Entry
"""
print("Initial Data Entry Step Start")
driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE&passurl&mntitle=Initial%20Data%20Entry&tc=1.0&stg=DE")

WebDriverWait(driver, 620).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# access element iframe
WebDriverWait(driver, 620).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_EKTP_NO']"))).send_keys(eKTPData)

WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_df_BARCODE']"))).send_keys(qrCodeData)

WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']"))).send_keys(branchData)
WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
time.sleep(5)

WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_df_SALES2']"))).click()
time.sleep(5)
WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_df_SALES2']"))).send_keys(salesIDData)
time.sleep(3)

selectChannel = Select(driver.find_element(By.ID, "mainPanel_channelPanel_df_CH_CODE"))
selectChannel.select_by_value(str(channelData))
time.sleep(3)
WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, "mainPanel_programPanel_df_PR_CODE")))
selectProgram = Select(driver.find_element(By.ID, "mainPanel_programPanel_df_PR_CODE"))
selectProgram.select_by_value(str(kodeProgramData))
time.sleep(3)
WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, "mainPanel_custcatPanel_df_CUSTCATID")))
selectJenisNasabah = Select(driver.find_element(By.ID, "mainPanel_custcatPanel_df_CUSTCATID"))
selectJenisNasabah.select_by_value(str(jenisNasabahData))
time.sleep(3)
WebDriverWait(driver, 60).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NMFIRST']").send_keys(namaNasabahData)

driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNDATE_DD']").send_keys(tanggalLahirData)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_BORNDATE_MM")))
selectBulanLahir = Select(driver.find_element(By.ID, "mainPanel_CU_BORNDATE_MM"))
selectBulanLahir.select_by_value(str(bulanLahirData))
time.sleep(3)

driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNDATE_YY']").send_keys(tahunLahirData)

driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_MMNMFIRST']").send_keys(namaIbuData)

driver.find_element(By.ID,"mainPanel_btn_save").click()
time.sleep(5)

# get Text Pop Up message
WebDriverWait(driver, 30).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert_text = alert.text
pattern = r'\bKNP\w+'
match = re.search(pattern, alert_text)
result = match.group(0)
workbook = load_workbook('Data\LOSData.xlsx')
sheet1 = workbook['Entry Data']
sheet1.cell(row=2, column=33).value = result
workbook.save('Data\LOSData.xlsx')
alert.accept()
print("Alert text:", result)

# exit from iframe 
driver.switch_to.default_content()
time.sleep(3)
print("Initial Data Entry Step End")

"""
Document Checking
"""
print("Document Checking Step Start")
noAplikasiData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 33)
print(noAplikasiData)
# Update Flag to 1.1 / Menu Document Checking
update_appflag(noAplikasiData, ap_currtrcode="1.1", ap_lasttrcode="1.0")

# Akses Menu Document Checking
driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE|D&passurl&mntitle=Document%20Checking&tc=1.1&regno={noAplikasiData}&stg=DE")

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen yang ada di dalam iframe
docCheckDataList = [docCheckData1, docCheckData2, docCheckData3]
for docCheckData in docCheckDataList:
    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "griddoc_header7_Button4"))).click()
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    update_document(driver, docCheckData)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    time.sleep(3)

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
time.sleep(3)
WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "updPanel_btnUpdate"))).click()

WebDriverWait(driver, 30).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert_text = alert.text
alert.accept()
WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
time.sleep(2)
WebDriverWait(driver, 30).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert_text = alert.text
alert.accept()

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

# exit from iframe 
driver.switch_to.default_content()

print("Document Checking Step End")

"""
Data Entry Assignment
"""
print("Data Entry Assignment Step Start")

# Akses Menu Data Entry Assignment
driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV|DE&passurl&mntitle=Data%20Entry%20Assignment&li1=LR|A|SPV|DE|00&li2=LR|A|SPV|DE|01&tc1=3.0&tc2=3.1&atype=DE&stg=DE&xlcode=DE")

WebDriverWait(driver, 320).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

assign_data(driver, "mainPanel_sendtoPanel_SendTo", noAplikasiData, sendToData)

print("Data Entry Assignment Step End")

"""
Data Entry
"""
print("Data Entry Step Start")

# Update User
update_user(noAplikasiData)
time.sleep(3)

# Akses Menu Data Entry
driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={noAplikasiData}&stg=DE")

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Personal Info']"))).click()

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen yang ada di dalam iframe
WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_CU_NICKNM']")))
driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NICKNM']").send_keys(namaPanggilanData)
driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NMCARD']").send_keys(namaCetakData)
driver.find_element(By.XPATH, f"(//input[@id='mainPanel_CU_GENDER_CODE_{str(jenisKelaminData)}'])[1]").click()
driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNPLACE']").send_keys(tempatLahirData)
driver.find_element(By.ID, "mainPanel_CU_EMAIL").send_keys(emailData)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_MR_CODE")))
selectStatusPernikahan = Select(driver.find_element(By.ID, "mainPanel_CU_MR_CODE"))
selectStatusPernikahan.select_by_value(str(statusPernikahanData))
driver.find_element(By.ID, "mainPanel_CU_NPWP").send_keys(NPWPData)
driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_CHILD']").send_keys(jumlahTanggunganData)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_EMAIL")))
selectPendidikanTerakhir = Select(driver.find_element(By.ID, "mainPanel_CU_ED_CODE"))
selectPendidikanTerakhir.select_by_value(str(pendidikanData))
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_TAXIDFLAG")))
selectIdentitasPajak = Select(driver.find_element(By.ID, "mainPanel_CU_TAXIDFLAG"))
selectIdentitasPajak.select_by_value(str(identitasPajakData))
driver.find_element(By.ID, "mainPanel_CU_HMADDR1").send_keys(alamatRumahData)
driver.find_element(By.ID, "mainPanel_CU_HMCITY").send_keys(kotaData)
driver.find_element(By.ID, "mainPanel_CU_HMZIPCODE_txtID").send_keys(kodePosData)
driver.find_element(By.ID, "mainPanel_CU_HMPHONEAREA").send_keys(telpRumahPrefixData)
driver.find_element(By.ID, "mainPanel_CU_HMPHONE").send_keys(telpRumahData)
driver.find_element(By.ID, "mainPanel_CU_HPNO").send_keys(telpHPData)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_HMSTA_CODE")))
selectStatusRumah = Select(driver.find_element(By.ID, "mainPanel_CU_HMSTA_CODE"))
selectStatusRumah.select_by_value(str(statusRumahData))
driver.find_element(By.ID, "mainPanel_CU_HMLIVESINCEYY").send_keys(lamaMenempatiThnData)
driver.find_element(By.ID, "mainPanel_CU_HMLIVESINCEMM").send_keys(lamaMenempatiBlnData)

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
WebDriverWait(driver, 10).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
time.sleep(3)
WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

# exit from iframe 
driver.switch_to.default_content()

WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Job Info']"))).click()

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen yang ada di dalam iframe
WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.ID, "mainPanel_CU_COMPNAME")))
driver.find_element(By.ID, "mainPanel_CU_COMPNAME").send_keys(namaPerusahaanData)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_JOB_CODE")))
selectStatusPernikahan = Select(driver.find_element(By.ID, "mainPanel_CU_JOB_CODE"))
selectStatusPernikahan.select_by_value(str(kategoriPekerjaanData))
driver.find_element(By.ID, "mainPanel_CU_BUSSINESSDESC").send_keys(bidangUsahaData)
driver.find_element(By.ID, "mainPanel_CU_WORKSINCEYY").send_keys(lamaBekerjaThnData)
driver.find_element(By.ID, "mainPanel_CU_WORKSINCEMM").send_keys(lamaBekerjaBlnData)
driver.find_element(By.ID, "mainPanel_CU_OFTITLE").send_keys(pangkatData)
driver.find_element(By.ID, "mainPanel_CU_INCOME").send_keys(pendapatanData)
driver.find_element(By.ID, "mainPanel_CU_OFADDR1").send_keys(alamatKantorData)
driver.find_element(By.ID, "mainPanel_CU_OFCITY").send_keys(kotaKantorData)
driver.find_element(By.ID, "mainPanel_CU_OFPHONEAREA").send_keys(telpKantorPrefixData)
driver.find_element(By.ID, "mainPanel_CU_OFPHONE").send_keys(telpKantorData)

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
WebDriverWait(driver, 10).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()

# exit from iframe 
driver.switch_to.default_content()

WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Emergency Contact']"))).click()

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen yang ada di dalam iframe
WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.ID, "mainPanel_CU_EMNMFIRST")))
driver.find_element(By.ID, "mainPanel_CU_EMNMFIRST").send_keys(namaEmergency)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_EMRE_CODE")))
selectHubuganEmergency = Select(driver.find_element(By.ID, "mainPanel_CU_EMRE_CODE"))
selectHubuganEmergency.select_by_value(str(hubunganEmergency))
driver.find_element(By.ID, "mainPanel_CU_EMHMADDR1").send_keys(alamatEmergency)
driver.find_element(By.ID, "mainPanel_CU_EMHMCITY").send_keys(kotaEmergency)
driver.find_element(By.ID, "mainPanel_CU_EMHMPHONEAREA").send_keys(telpRumahPrefixEmergency)
driver.find_element(By.ID, "mainPanel_CU_EMHMPHONE").send_keys(telpRumahEmergency)

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
WebDriverWait(driver, 10).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
time.sleep(3)
WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

# exit from iframe 
driver.switch_to.default_content()

WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Other Options']"))).click()

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen yang ada di dalam iframe
WebDriverWait(driver, 180).until(EC.presence_of_element_located((By.ID, f"mainPanel_BILLING_DLV_TYPE_0")))
driver.find_element(By.ID, f"mainPanel_BILLING_DLV_TYPE_{tagihanData}").click()
driver.find_element(By.ID, f"mainPanel_MAINCARD_DLV_TYPE_{kartuUtamaData}").click()
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_BILLING_STAT_TYPE")))
selectPilihanBS = Select(driver.find_element(By.ID, "mainPanel_BILLING_STAT_TYPE"))
selectPilihanBS.select_by_value(str(pilihanBSData))
driver.find_element(By.ID, "mainPanel_AD_ACCNUM").send_keys('0')

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
time.sleep(3)
WebDriverWait(driver, 50).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
time.sleep(3)
WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

# exit from iframe 
driver.switch_to.default_content()

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Main']"))).click()

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen yang ada di dalam iframe
WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "mainPanel_GV_PRODUCT_header5_BTN_NEW"))).click()
time.sleep(3)

WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='popupDetail_panelDetail_df_NETWORKID']")))
selectNetworkCard = Select(driver.find_element(By.XPATH, "//*[@id='popupDetail_panelDetail_df_NETWORKID']"))
selectNetworkCard.select_by_value(networkData)
time.sleep(3)

WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "popupDetail_panelDetail_df_PRODUCTID")))
selectProductCard = Select(driver.find_element(By.ID, "popupDetail_panelDetail_df_PRODUCTID"))
selectProductCard.select_by_value(productData)
time.sleep(3)

if (networkData in ["G", "J", "P"] and productData == "P") or \
   (networkData == "M" and productData in ["C", "G", "P"]) or \
   (networkData == "V"):
    select_dropdown_by_value(driver, "popupDetail_panelDetail_cardtypePanel_df_CARDTYPEID", cardTypeData)

time.sleep(3)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "popupDetail_panelDetail_plastictypePanel_df_PLASTICID")))
selectProductCard = Select(driver.find_element(By.ID, "popupDetail_panelDetail_plastictypePanel_df_PLASTICID"))
selectProductCard.select_by_value(str(plasticTypeData))

WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "popupDetail_panelDetail_Button3"))).click()
time.sleep(3)

WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.ID, 'popupDetail_CSD-1')))

WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
WebDriverWait(driver, 620).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
time.sleep(3)
WebDriverWait(driver, 620).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
time.sleep(3)
WebDriverWait(driver, 620).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
driver.switch_to.default_content()
print("Data Entry Step End")

"""
DE Validation Assignment
"""
print("DE Validation Assignment Step Start")
driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DE%20Validation%20Assignment&li1=LR|SPV|DE|00&li2=LR|SPV|01&tc1=3.5.0&tc2=3.5.1&stg=VDE")

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

assign_data(driver, "mainPanel_SendTo", noAplikasiData, sendToData)

print("DE Validation Assignment Step End")

"""
DE Validation
"""
time.sleep(3)
print("Data Entry Validation Step Start")

# Update User
update_user(noAplikasiData)
print("sukses update user")
time.sleep(3)

driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VDE&passurl&mntitle=Data%20Entry%20Validation&tc=3.5.1&regno={noAplikasiData}&stg=DE")
time.sleep(3)

# Access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# Access elemen iframe
WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
WebDriverWait(driver, 50).until(EC.alert_is_present())
alert = driver.switch_to.alert
print(alert.text)
alert.accept()  
time.sleep(5)
WebDriverWait(driver, 50).until(EC.alert_is_present())
alert = driver.switch_to.alert
print(alert.text)
alert.accept()
time.sleep(3)
WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

# Exit from iframe 
driver.switch_to.default_content()

print("Data Entry Validation Step End")

# """
# Scoring
# """
# time.sleep(3)
# host = '172.24.141.33'
# username = 'administrator'
# password = 'c@rdl1nk'
# pathWinSCP = r'C:/Program Files (x86)/WinSCP/WinSCP.exe'

# appFlagAfter = automate_scoring(pathWinSCP, noAplikasiData, host, username, password)
# print("Proses Scoring Selesai")
# print("Flag after proses scoring", appFlagAfter)

# if appFlagAfter in ["9.2.1", "9.2.3"]:
#     sys.exit(f"Program dihentikan karena appFlagAfter {appFlagAfter}")
# elif appFlagAfter in ["3.5.0", "3.6.1"]:
#     while appFlagAfter in ["3.5.0", "3.6.1"]:
#         print(f"Step Scoring Diulang karena flag {appFlagAfter}")
#         time.sleep(5)
#         appFlagAfter = automate_scoring(pathWinSCP, noAplikasiData, host, username, password)
#         print("Proses Scoring Selesai")
#         print("Flag after proses scoring", appFlagAfter)

"""
Credit Reviewer Assignment
"""
print("Credit Reviewer Assignment Start")

# Akses Menu Credit Reviewer Assignment
time.sleep(3)
driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=Credit%20Reviewer%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=5.7.0&tc2=5.7.1&stg=ANLMINI")
time.sleep(3)

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

assign_data(driver, "mainPanel_SendTo", noAplikasiData, sendToAppReviewer)

print("Credit Reviewer Assignment End")

"""
Credit Reviewer
"""
print("Credit Reviewer  Start")

# Update User
update_user(noAplikasiData)
time.sleep(3)

# Akses Menu Credit Reviewer 
driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=ANLMINI&passurl&mntitle=Credit%20Reviewer&tc=5.7.1&regno={noAplikasiData}&stg=ANL")

time.sleep(10)
WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Decision']"))).click()
# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

# akses elemen iframe
WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_GV_PRODUCT_cell0_6_BTN_EDIT"))).click()
time.sleep(3)
WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
WebDriverWait(driver, 320).until(EC.visibility_of_element_located((By.ID, "popupDetail_panelDetail_df_CP_APRLIMIT"))).send_keys(str(limitData))
time.sleep(3)
WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "popupDetail_panelDetail_Button6"))).click()
time.sleep(3)

WebDriverWait(driver, 620).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
WebDriverWait(driver, 620).until(EC.invisibility_of_element_located((By.ID, 'popupDetail_CSD-1')))
time.sleep(3)

WebDriverWait(driver, 320).until(EC.presence_of_element_located((By.ID, 'mainPanel_decPanel_dfa_AP_NEXTTRBY')))
select_element = Select(driver.find_element(By.ID, 'mainPanel_decPanel_dfa_AP_NEXTTRBY'))
select_element.select_by_value("DIRTONI")
time.sleep(3)

WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
time.sleep(3)


WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_decPanel_btn_update']"))).click()
WebDriverWait(driver, 60).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
time.sleep(3)
WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

# exit from iframe 
driver.switch_to.default_content()

print("Credit Reviewer  Step End")

"""
Approval
"""
print("Approval Step Start")
time.sleep(5)

# Update User
update_user(noAplikasiData, ap_nexttrby='cc_dam')
print("update user berhasil")
time.sleep(5)
update_user(noAplikasiData, ap_nexttrby='cc_dam')
print("update user berhasil")

# akses menu approval
driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=APRV&passurl&mntitle=Approval&tc=7.1&regno={noAplikasiData}&stg=APRV")
time.sleep(10)
WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Decision']"))).click()

# access iframe
WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_decPanel_btn_appr"))).click()
WebDriverWait(driver, 60).until(EC.alert_is_present())
alert = driver.switch_to.alert
alert.accept()
WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
time.sleep(5)

check81 = check_appflag(str(noAplikasiData))
pattern = re.compile(r"'(.*?)'")
is81 = pattern.search(str(check81))

if (is81.group(1) == '7.1') :
    while is81.group(1) == '7.1':
        print(f"Proses Approval Diulang karena flag {is81.group(1)}")
        print("Approval Step Start")
        time.sleep(5)
        # akses menu approval
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=APRV&passurl&mntitle=Approval&tc=7.1&regno={noAplikasiData}&stg=APRV")
        time.sleep(5)

        driver.find_element(By.XPATH, "//a[normalize-space()='Decision']").click()
        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_decPanel_btn_appr"))).click()
        WebDriverWait(driver, 30).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(5)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

print("Approval Step End")

# exit from iframe 
driver.switch_to.default_content()
time.sleep(10)

if (noRekData != 0):
    print("Update Norek Data")
    time.sleep(3)
    update_norek(str(noAplikasiData), str(noRekData))
    time.sleep(3)
    print("Update Norek Data Selesai")

print("Proses Onboarding Data LOS Selesai")
elapsed_seconds = int(time.time() - start_time)
print(f"Proses berjalan selama : {elapsed_seconds} detik")
driver.quit()