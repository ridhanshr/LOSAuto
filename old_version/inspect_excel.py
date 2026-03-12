from openpyxl import load_workbook
import os

file_path = 'Data/LOSData.xlsx'
output_path = 'excel_info.txt'
with open(output_path, 'w') as f:
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        f.write(f"Sheets: {wb.sheetnames}\n")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = []
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=1, column=c).value
                if val:
                    headers.append(str(val))
                else:
                    break
            f.write(f"Sheet '{sheet_name}' headers: {', '.join(headers)}\n")
    else:
        f.write("File not found.\n")
print("Done")
