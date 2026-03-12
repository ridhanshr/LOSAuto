from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from src.utils.db_utils import update_appflag, update_user, update_accnum, check_appflag, update_applicant, update_norek
from src.utils.login_utils import perform_login
from src.utils.xlsx_utils import read_excel_data, count_filled_rows
from src.utils.docCheck_utils import update_document
from src.utils.dataEntry_utils import select_dropdown_by_value
from src.utils.assign_utils import assign_data
from src.config import config

import time
import datetime
import re
import sys
import random
import os
from faker import Faker

"""
setup webdriver
"""
def setup_webdriver(browser_type="edge"):
    """Setup webdriver based on configuration"""
    if browser_type.lower() == "chrome":
        chrome_path = r'drivers/chromedriver.exe'
        service = Service(executable_path=chrome_path)
        chrome_options = ChromeOptions()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--remote-allow-origins=*")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-software-rasterizer")
        if config.get("headless", False):
            chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=service, options=chrome_options)
    else:  # Default to Edge
        edge_path = r'drivers/msedgedriver.exe'
        service = Service(executable_path=edge_path)
        edge_options = EdgeOptions()
        edge_options.add_argument("--log-level=1")
        edge_options.add_argument("--remote-allow-origins=*")
        if config.get("headless", False):
            edge_options.add_argument("--headless")
        driver = webdriver.Edge(service=service, options=edge_options)
    
    driver.maximize_window()
    return driver

def save_screenshot(driver, filename, screenshot_dir=None):
    """Save screenshot to configured directory"""
    if screenshot_dir is None:
        screenshot_dir = config.get("screenshot_dir", "Data/screenshoot")
    
    # Ensure directory exists
    os.makedirs(screenshot_dir, exist_ok=True)
    
    # Save screenshot
    screenshot_path = os.path.join(screenshot_dir, filename)
    driver.save_screenshot(screenshot_path)
    print(f"Screenshot saved: {screenshot_path}")

def main():
    # Parse command line arguments
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser", default=config.get("browser", "edge"), help="Browser type")
    args, unknown = parser.parse_known_args()

    # Setup webdriver based on argument or config
    driver = setup_webdriver(args.browser)

    """
    setup data
    """
    # Setup path data from config
    xlsxdataPath = config.get("data_file", "Data/LOSData.xlsx")
    sheetEntryData = "Entry Data"
    sheetEntryDataAddOn = "Entry Data Add On"
    sheetSupplement = "Supplement"
    sheetDocCheck = "Document Checking"
    sheetAssignment = "Assignment"
    sheetCard = "Card"

    # Retrieve Data
    # Data Entry Data
    fake = Faker('en_NZ')
    qrCodeData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 1)
    branchData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 2)
    channelData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 3)
    kodeProgramData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 4)
    jenisNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 5)

    # Entry Data Add On
    noKartu = read_excel_data(xlsxdataPath, sheetEntryDataAddOn, 2, 1)

    # Document Check Data
    docCheckData1 = read_excel_data(xlsxdataPath, sheetDocCheck, 2, 1)
    docCheckData2 = read_excel_data(xlsxdataPath, sheetDocCheck, 3, 1)
    docCheckData3 = read_excel_data(xlsxdataPath, sheetDocCheck, 4, 1)
    docCheckData4 = read_excel_data(xlsxdataPath, sheetDocCheck, 5, 1)
    docCheckData5 = read_excel_data(xlsxdataPath, sheetDocCheck, 6, 1)

    #Card Data
    networkData = read_excel_data(xlsxdataPath, sheetCard, 2, 1)
    productData = read_excel_data(xlsxdataPath, sheetCard, 2, 2)
    cardTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 3)
    plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 4)
    limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)
    namaNasabahAddOn = fake.name_male()
    plasticTypeAddOnData = read_excel_data(xlsxdataPath, sheetCard, 2, 7)
    workbook = load_workbook(xlsxdataPath)
    sheet1 = workbook['Card']
    sheet1.cell(row=2, column=6).value = namaNasabahAddOn
    workbook.save(xlsxdataPath)

    # Assignment Data
    sendToData = read_excel_data(xlsxdataPath, sheetAssignment, 2, 1)
    sendToAppReviewer = read_excel_data(xlsxdataPath, sheetAssignment, 2, 2)

    start_time = time.time()
    """
    login
    """
    # access los
    driver.get("http://172.24.141.61/bricc/unittest.aspx")
    print("Login")
    credentials = config.get_credentials()
    username = credentials.get("username", "cc_dam")
    perform_login(driver, username)

    """
    generate QR
    """
    time.sleep(10)
    print("Generate QR Step Start")
    driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=BAR|G&passurl&mntitle=Generate%20Barcode&row=8&barfontsize=40&barsize=100&imgwidth=250&borderwidth=0&altalign=1&padtop=20&rowheight=73&colwidth=350&cellspacing=0")

    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    # access iframe
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

    # access element iframe
    driver.find_element(By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']").send_keys(branchData)
    driver.find_element(By.XPATH, "//input[@id='mainPanel_btn_gen']").click()
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    image_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='mainPanel_tbl']/tbody/tr[2]/td[1]/img")))
    image_url = image_element.get_attribute('src')
    qrCode = image_url.split("data=")[1].split("&")[0]
    qrCodeSub = re.sub(r'^(.....)1(.*)$', r'\g<1>0\g<2>', qrCode)
    workbook = load_workbook(xlsxdataPath)
    sheet1 = workbook['Entry Data']
    sheet1.cell(row=2, column=1).value = qrCodeSub
    workbook.save(xlsxdataPath)

    # exit from iframe 
    driver.switch_to.default_content()

    print("Generate QR Step End")

    """
    initial data entry add on
    """
    print("initial data add on")
    driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE|A&passurl&mntitle=Add-On%20Supplement&tc=1.0&stg=DE")
    time.sleep(1)

    # access iframe
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "(//input[@name='mainPanel$ctl04'])[1]"))).click()
    time.sleep(1)
    driver.find_element(By.XPATH, "//input[@id='PopFindBasic_PNL_FindBasic_fCORE_ID']").send_keys(noKartu)
    time.sleep(2)
    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Find']"))).click()
    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "PopFindBasic_PNL_FindBasic_GridViewX_cell0_8_Button4"))).click()
    time.sleep(3)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    driver.find_element(By.CSS_SELECTOR, "#mainPanel_branchPanel_df_BRANCHID_txtID").clear()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, "#mainPanel_branchPanel_df_BRANCHID_txtID").send_keys(branchData)
    time.sleep(5)
    driver.find_element(By.XPATH, "//input[@id='mainPanel_df_BARCODE']").send_keys(qrCodeSub)
    time.sleep(10)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    WebDriverWait(driver, 9999999).until(EC.visibility_of_element_located((By.XPATH, "//select[@id='mainPanel_channelPanel_df_CH_CODE']"))).send_keys(channelData)
    time.sleep(10)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    WebDriverWait(driver, 9999999).until(EC.visibility_of_element_located((By.XPATH, "//select[@id='mainPanel_programPanel_df_PR_CODE']"))).send_keys(kodeProgramData)
    time.sleep(10)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    WebDriverWait(driver, 9999999).until(EC.visibility_of_element_located((By.XPATH, "//select[@id='mainPanel_custcatPanel_df_CUSTCATID']"))).send_keys(jenisNasabahData)
    time.sleep(10)

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_Button2']"))).click()
    time.sleep(3)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()  
    time.sleep(5)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()
    time.sleep(3)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

    # Exit from iframe 
    driver.switch_to.default_content()

    """
    Document Checking
    """
    print("Document Checking Step Start")
    time.sleep(2)
    # ambil no aplikasi
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
    elementNoAplikasi = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "UC_GeneralInfo1_l1")))
    valueNoAplikasi = elementNoAplikasi.text
    time.sleep(2)
    workbook = load_workbook(xlsxdataPath)
    sheet1 = workbook['Entry Data']
    sheet1.cell(row=2, column=39).value = valueNoAplikasi
    workbook.save(xlsxdataPath)

    # akses elemen yang ada di dalam iframe
    docCheckDataList = [docCheckData1, docCheckData3]
    for docCheckData in docCheckDataList:
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "griddoc_header7_Button4"))).click()
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        update_document(driver, docCheckData)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(3)

    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    time.sleep(3)
    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "updPanel_btnUpdate"))).click()

    WebDriverWait(driver, 30).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    alert_text = alert.text
    alert.accept()
    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
    time.sleep(2)
    WebDriverWait(driver, 30).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    alert_text = alert.text
    alert.accept()

    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    # exit from iframe 
    driver.switch_to.default_content()

    print("Document Checking Step End")

    """
    Data Entry Assignment
    """
    print("Data Entry Assignment Step Start")

    # Akses Menu Data Entry Assignment
    driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV|DE&passurl&mntitle=Data%20Entry%20Assignment&li1=LR|A|SPV|DE|00&li2=LR|A|SPV|DE|01&tc1=3.0&tc2=3.1&atype=DE&stg=DE&xlcode=DE")

    WebDriverWait(driver, 320).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    assign_data(driver, "mainPanel_sendtoPanel_SendTo", valueNoAplikasi, sendToData)

    print("Data Entry Assignment Step End")

    """
    Data Entry
    """
    print("Data Entry Step Start")

    # Update User
    update_user(valueNoAplikasi)
    time.sleep(3)

    # Akses Menu Data Entry
    driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={valueNoAplikasi}&stg=DE")

    WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Supplement']"))).click()

    # access iframe
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
    time.sleep(1)

    # Supplement Data
    namaSupplement = fake.name_male() + "BRIMO"
    print(namaSupplement)
    tanngalLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 2)
    bulanLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 3)
    print("bulanLahirSupplement", bulanLahirSupplement)
    tahunLahirSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 4)
    noHPSupplement = read_excel_data(xlsxdataPath, sheetSupplement, 2, 8)

    workbook = load_workbook(xlsxdataPath)
    sheet1 = workbook['Supplement']
    sheet1.cell(row=1, column=2).value = namaSupplement
    workbook.save(xlsxdataPath)

    # akses elemen yang ada di dalam iframe
    WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "GridViewSuppInfo_header6_Button5"))).click()
    time.sleep(3)
    driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_NAME").send_keys(namaSupplement)
    driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_DD").send_keys(bulanLahirSupplement)
    WebDriverWait(driver, 180).until(EC.presence_of_element_located((By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_MM")))
    selectBulanLahirSupplement = Select(driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_MM"))
    selectBulanLahirSupplement.select_by_value(str(bulanLahirSupplement))
    time.sleep(3)
    driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_BORNDATE_YY").send_keys(tahunLahirSupplement)
    driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_LIMIT").send_keys(limitData)
    driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_HPNO").send_keys(noHPSupplement)
    driver.find_element(By.ID, "PopupSuppInfo_PNL_SuppInfo_SUP_NMCARD").send_keys(namaSupplement)

    WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "PopupSuppInfo_PNL_SuppInfo_BTN_SAVE"))).click()
    time.sleep(3)

    # exit from iframe 
    driver.switch_to.default_content()

    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Main']"))).click()

    update_accnum(valueNoAplikasi)
    time.sleep(1)

    # access iframe
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
    WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
    WebDriverWait(driver, 620).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    alert.accept()
    time.sleep(3)
    WebDriverWait(driver, 620).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    alert.accept()
    WebDriverWait(driver, 620).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
    print("Data Entry Step End")

    # Exit from iframe 
    driver.switch_to.default_content()

    """
    DE Validation Assignment
    """
    print("DE Validation Assignment Step Start")
    driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DE%20Validation%20Assignment&li1=LR|SPV|DE|00&li2=LR|SPV|01&tc1=3.5.0&tc2=3.5.1&stg=VDE")

    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, sendToData)

    print("DE Validation Assignment Step End")

    """
    DE Validation
    """
    time.sleep(3)
    print("Data Entry Validation Step Start")

    # Update User
    update_user(valueNoAplikasi)
    print("sukses update user")
    time.sleep(3)

    driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VDE&passurl&mntitle=Data%20Entry%20Validation&tc=3.5.1&regno={valueNoAplikasi}&stg=DE")
    time.sleep(3)

    # Access iframe
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

    # Access elemen iframe
    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()  
    time.sleep(5)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()
    time.sleep(3)
    WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

    # Exit from iframe 
    driver.switch_to.default_content()

    print("Data Entry Validation Step End")

    """
    Verphone Add Basic Supplement Assignment
    """
    update_appflag(valueNoAplikasi, '5.5.0', '3.5.1')
    time.sleep(3)

    print("Verphone Add Basic Supplement Assignment Start")
    driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=VerPhone%20Supplement%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=5.5.0&tc2=5.5.1&stg=VERSUP")

    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, 'MY_BRI_APPROVAL')

    print("Verphone Add Basic Supplement Assignment End")

    """
    Verphone Supplement
    """
    # Update User
    update_user(valueNoAplikasi)
    print("sukses update user")
    time.sleep(3)

    print("Verphone Supplement Start")
    driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VERSUP&passurl&mntitle=VerPhone%20Supplement&tc=5.5.1&regno={valueNoAplikasi}")

    # access iframe
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_Button6"))).click()
    time.sleep(3)

    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()  
    time.sleep(5)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()
    time.sleep(3)

    # exit from iframe 
    driver.switch_to.default_content()
    print("Verphone Supplement End")

    """
    Data Maintenance Assignment
    """
    print("Data Maintenance Assignment Start")
    driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=Data%20Maintenance%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=6.5.0&tc2=6.5.1&stg=DMSUP")

    WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

    assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, 'MY_BRI_APPROVAL')

    print("Data Maintenance Assignment End")

    """
    Data Maintenance for Suplement
    """
    # Update User
    update_user(valueNoAplikasi)
    print("sukses update user")
    time.sleep(3)

    print("Data Maintenance for Suplement Start")
    driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DMSUP&passurl&mntitle=Data%20Maintenance%20for%20Supplement&tc=6.5.1&regno={valueNoAplikasi}")
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
    time.sleep(1)

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "GridViewSuppInfo_cell0_6_Button1"))).click()
    time.sleep(3)

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, f"PopupSuppInfo_PNL_SuppInfo_SUP_PLASTICID_{str(plasticTypeAddOnData)}"))).click()
    dateTimeNow = datetime.datetime.now()
    driver.find_element(By.ID, 'PopupSuppInfo_PNL_SuppInfo_SUP_POT_EXPMM').send_keys("0" + str(dateTimeNow.month))
    driver.find_element(By.ID, 'PopupSuppInfo_PNL_SuppInfo_SUP_POT_EXPYY').send_keys(str(dateTimeNow.year + 5))
    time.sleep(2)
    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "PopupSuppInfo_PNL_SuppInfo_Button1"))).click()

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_btn_appr"))).click()
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()  
    time.sleep(5)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()
    time.sleep(1)
    print("Data Maintenance for Suplement End")
    # exit from iframe 
    driver.switch_to.default_content()

    """
    Data Maintenance for Supplement Approval
    """
    print("Data Maintenance for Supplement Approval Start")
    driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DMSUP&passurl&mntitle=Data%20Maintenance%20for%20Supplement%20Approval&tc=6.5.2&regno={valueNoAplikasi}")
    WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
    time.sleep(1)

    WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_btn_appr"))).click()
    time.sleep(3)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()  
    time.sleep(5)
    WebDriverWait(driver, 50).until(EC.alert_is_present())
    alert = driver.switch_to.alert
    print(alert.text)
    alert.accept()
    time.sleep(1)
    print("Data Maintenance for Supplement Approval End")

    # exit from iframe 
    driver.switch_to.default_content()

    print("Proses Onboarding Data Add On Basic LOS Selesai")
    elapsed_seconds = int(time.time() - start_time)
    print(f"Proses berjalan selama : {elapsed_seconds} detik")
    driver.quit()

if __name__ == "__main__":
    main()