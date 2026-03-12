from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook

# Local utilities
from src.utils.db_utils import update_appflag, update_user, update_accnum
from src.utils.login_utils import perform_login
from src.utils.xlsx_utils import read_excel_data, count_filled_rows
from src.utils.docCheck_utils import update_document
from src.utils.dataEntry_utils import select_dropdown_by_value
from src.utils.assign_utils import assign_data
from src.utils.ui_utils import setup_webdriver, save_screenshot
from src.config import config

# Standard libraries
import time
import re
import logging

from faker import Faker
from src.utils.logging_utils import setup_logging


def main():
    # ========== SETUP ==========
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser", default=config.get("browser", "edge"), help="Browser type")
    parser.add_argument("--log-level", default=config.get("log_level", "info"), help="Log level")
    args, unknown = parser.parse_known_args()

    # Setup logging
    setup_logging(args.log_level)
    logger = logging.getLogger("LOSAuto.AddOnLOS")

    # Setup webdriver based on argument or config
    driver = setup_webdriver(args.browser)

    # ========== DATA CONFIGURATION ==========
    xlsxdataPath = config.get("data_file", "Data/LOSData.xlsx")
    sheetEntryData = "Entry Data"
    sheetEntryDataAddOn = "Entry Data Add On"
    sheetDocCheck = "Document Checking"
    sheetAssignment = "Assignment"
    sheetCard = "Card"

    # Initialize Faker
    fake = Faker('en_NZ')

    # ========== READ EXCEL DATA ==========
    # Entry Data
    qrCodeData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 1)
    branchData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 2)
    channelData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 3)
    kodeProgramData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 4)
    jenisNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 5)

    # Entry Data Add On
    noKartu = read_excel_data(xlsxdataPath, sheetEntryDataAddOn, 2, 1)
    logger.info(f"No Kartu: {noKartu}")

    # Document Check Data
    docCheckData1 = read_excel_data(xlsxdataPath, sheetDocCheck, 2, 1)
    docCheckData2 = read_excel_data(xlsxdataPath, sheetDocCheck, 3, 1)
    docCheckData3 = read_excel_data(xlsxdataPath, sheetDocCheck, 4, 1)
    docCheckData4 = read_excel_data(xlsxdataPath, sheetDocCheck, 5, 1)
    docCheckData5 = read_excel_data(xlsxdataPath, sheetDocCheck, 6, 1)

    # Card Data
    networkData = read_excel_data(xlsxdataPath, sheetCard, 2, 1)
    productData = read_excel_data(xlsxdataPath, sheetCard, 2, 2)
    cardTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 3)
    plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 4)
    limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)
    namaNasabahAddOn = fake.name_male()
    plasticTypeAddOnData = read_excel_data(xlsxdataPath, sheetCard, 2, 7)
    
    # Save generated name to Excel
    workbook = load_workbook(xlsxdataPath)
    sheet1 = workbook['Card']
    sheet1.cell(row=2, column=6).value = namaNasabahAddOn
    workbook.save(xlsxdataPath)
    workbook.close()

    # Assignment Data
    sendToData = read_excel_data(xlsxdataPath, sheetAssignment, 2, 1)
    sendToAppReviewer = read_excel_data(xlsxdataPath, sheetAssignment, 2, 2)

    try:
        start_time = time.time()

        # ========== LOGIN ==========
        logger.info("Login Step Start")
        driver.get("http://172.24.141.61/bricc/unittest.aspx")
        credentials = config.get_credentials()
        username = credentials.get("username", "cc_dam")
        perform_login(driver, username)
        logger.info("Login Step End")

        # ========== GENERATE QR ==========
        logger.info("Generate QR Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=BAR|G&passurl&mntitle=Generate%20Barcode&row=8&barfontsize=40&barsize=100&imgwidth=250&borderwidth=0&altalign=1&padtop=20&rowheight=73&colwidth=350&cellspacing=0")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']").send_keys(branchData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_btn_gen']").click()
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        image_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='mainPanel_tbl']/tbody/tr[2]/td[1]/img")))
        image_url = image_element.get_attribute('src')
        save_screenshot(driver, '2.png')

        qrCode = image_url.split("data=")[1].split("&")[0]
        qrCodeSub = re.sub(r'^(.....)1(.*)$', r'\g<1>0\g<2>', qrCode)
        
        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=1).value = qrCodeSub
        workbook.save(xlsxdataPath)
        workbook.close()
        
        driver.switch_to.default_content()
        logger.info("Generate QR Step End")

        # ========== INITIAL DATA ENTRY ADD ON ==========
        logger.info("Initial Data Entry Add-On Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE|B&passurl&mntitle=Add-On%20Basic&tc=1.0&stg=DE")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[3]/div[3]/table/tbody/tr[9]/td[1]/input"))).click()
        time.sleep(1)
        
        driver.find_element(By.XPATH, "//input[@id='PopFindBasic_PNL_FindBasic_fCORE_ID']").send_keys(noKartu)
        time.sleep(2)
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Find']"))).click()
        time.sleep(2)
        save_screenshot(driver, '3.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "PopFindBasic_PNL_FindBasic_GridViewX_cell0_8_Button4"))).click()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        
        driver.find_element(By.CSS_SELECTOR, "#mainPanel_branchPanel_df_BRANCHID_txtID").clear()
        time.sleep(5)
        driver.find_element(By.CSS_SELECTOR, "#mainPanel_branchPanel_df_BRANCHID_txtID").send_keys(branchData)
        time.sleep(5)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_df_BARCODE']").send_keys(qrCodeSub)
        time.sleep(15)
        driver.find_element(By.XPATH, "//select[@id='mainPanel_channelPanel_df_CH_CODE']").send_keys(channelData)
        time.sleep(5)
        driver.find_element(By.XPATH, "//select[@id='mainPanel_programPanel_df_PR_CODE']").send_keys(kodeProgramData)
        time.sleep(5)
        driver.find_element(By.XPATH, "//select[@id='mainPanel_custcatPanel_df_CUSTCATID']").send_keys(jenisNasabahData)
        time.sleep(5)
        save_screenshot(driver, '4.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_Button2']"))).click()
        time.sleep(3)
        
        # Handle alerts
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                logger.info(f"Alert: {alert.text}")
                alert.accept()
                time.sleep(3)
            except:
                pass

        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        driver.switch_to.default_content()
        logger.info("Initial Data Entry Add-On Step End")

        # ========== DOCUMENT CHECKING ==========
        logger.info("Document Checking Step Start")
        time.sleep(2)
        
        # Get application number
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        elementNoAplikasi = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "UC_GeneralInfo1_l1")))
        valueNoAplikasi = elementNoAplikasi.text
        time.sleep(2)
        
        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=39).value = valueNoAplikasi
        workbook.save(xlsxdataPath)
        workbook.close()

        # Process documents
        docCheckDataList = [docCheckData1, docCheckData2, docCheckData3, docCheckData4, docCheckData5]
        for docCheckData in docCheckDataList:
            WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "griddoc_header7_Button4"))).click()
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            update_document(driver, docCheckData)
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            time.sleep(3)

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(3)
        save_screenshot(driver, '5.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "updPanel_btnUpdate"))).click()
        
        # Handle alerts
        for _ in range(2):
            try:
                WebDriverWait(driver, 30).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
                time.sleep(2)
            except:
                pass

        driver.switch_to.default_content()
        logger.info("Document Checking Step End")

        # ========== DATA ENTRY ASSIGNMENT ==========
        logger.info("Data Entry Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV|DE&passurl&mntitle=Data%20Entry%20Assignment&li1=LR|A|SPV|DE|00&li2=LR|A|SPV|DE|01&tc1=3.0&tc2=3.1&atype=DE&stg=DE&xlcode=DE")
        WebDriverWait(driver, 320).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_sendtoPanel_SendTo", valueNoAplikasi, sendToData)
        save_screenshot(driver, '6.png')
        logger.info("Data Entry Assignment Step End")

        # ========== DATA ENTRY ==========
        logger.info("Data Entry Step Start")
        update_user(valueNoAplikasi)
        time.sleep(3)

        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={valueNoAplikasi}&stg=DE")
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Other Options']"))).click()
        time.sleep(3)
        save_screenshot(driver, '7.png')

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Main']"))).click()
        time.sleep(3)
        
        workbook = load_workbook(xlsxdataPath)
        sheet = workbook['Card']
        filled_rows_card = count_filled_rows(sheet)
        workbook.close()

        # Add cards
        for row in range(2, filled_rows_card + 1):
            networkData = read_excel_data(xlsxdataPath, sheetCard, row, 1)
            productData = read_excel_data(xlsxdataPath, sheetCard, row, 2)
            cardTypeData = read_excel_data(xlsxdataPath, sheetCard, row, 3)
            plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, row, 4)

            time.sleep(3)
            WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
            WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "mainPanel_GV_PRODUCT_header5_BTN_NEW"))).click()
            time.sleep(3)

            WebDriverWait(driver, 180).until(EC.presence_of_element_located((By.XPATH, "//*[@id='popupDetail_panelDetail_df_NETWORKID']")))
            Select(driver.find_element(By.XPATH, "//*[@id='popupDetail_panelDetail_df_NETWORKID']")).select_by_value(networkData)
            time.sleep(3)

            WebDriverWait(driver, 180).until(EC.presence_of_element_located((By.ID, "popupDetail_panelDetail_df_PRODUCTID")))
            Select(driver.find_element(By.ID, "popupDetail_panelDetail_df_PRODUCTID")).select_by_value(productData)
            time.sleep(3)

            if (networkData in ["G", "J", "P"] and productData == "P") or \
               (networkData == "M" and productData in ["C", "G", "P"]) or \
               (networkData == "V"):
                select_dropdown_by_value(driver, "popupDetail_panelDetail_cardtypePanel_df_CARDTYPEID", cardTypeData)

            time.sleep(3)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "popupDetail_panelDetail_plastictypePanel_df_PLASTICID")))
            Select(driver.find_element(By.ID, "popupDetail_panelDetail_plastictypePanel_df_PLASTICID")).select_by_value(str(4))

            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "popupDetail_panelDetail_Button3"))).click()
            time.sleep(3)
            
            WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
            WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.ID, 'popupDetail_CSD-1')))

            driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={valueNoAplikasi}&stg=DE")
            driver.refresh()
            time.sleep(3)

        save_screenshot(driver, '8.png')

        # Submit Data Entry
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        
        for _ in range(2):
            try:
                WebDriverWait(driver, 620).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(3)
            except:
                pass
        
        WebDriverWait(driver, 620).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        driver.switch_to.default_content()
        logger.info("Data Entry Step End")

        # ========== DE VALIDATION ASSIGNMENT ==========
        logger.info("DE Validation Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DE%20Validation%20Assignment&li1=LR|SPV|DE|00&li2=LR|SPV|01&tc1=3.5.0&tc2=3.5.1&stg=VDE")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, sendToData)
        save_screenshot(driver, '9.png')
        logger.info("DE Validation Assignment Step End")

        # ========== DE VALIDATION ==========
        logger.info("Data Entry Validation Step Start")
        time.sleep(3)
        update_user(valueNoAplikasi)
        logger.info("Update user successful")
        time.sleep(3)

        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VDE&passurl&mntitle=Data%20Entry%20Validation&tc=3.5.1&regno={valueNoAplikasi}&stg=DE")
        time.sleep(3)
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        save_screenshot(driver, '10.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        
        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                logger.info(f"Alert: {alert.text}")
                alert.accept()
                time.sleep(3)
            except:
                pass

        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        driver.switch_to.default_content()
        logger.info("Data Entry Validation Step End")

        # ========== VERPHONE ADD BASIC ASSIGNMENT ==========
        update_appflag(valueNoAplikasi, '5.6.0', '3.5.1')
        time.sleep(3)

        logger.info("Verphone Add Basic Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=VerPhone%20Add%20Basic%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=5.6.0&tc2=5.6.1&stg=VERBAS")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, 'MY_BRI_APPROVAL')
        save_screenshot(driver, '11.png')
        logger.info("Verphone Add Basic Assignment Step End")

        # ========== VERPHONE ADD BASIC ==========
        logger.info("Verphone Add Basic Step Start")
        update_user(valueNoAplikasi)
        logger.info("Update user successful")
        time.sleep(3)

        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VERBAS&passurl&mntitle=VerPhone%20Add%20Basic&tc=5.6.1&regno={valueNoAplikasi}")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, f"(//input[@id='UpdPanel_df_PLASTICID_{plasticTypeAddOnData}'])[1]"))).click()
        time.sleep(1)
        WebDriverWait(driver, 620).until(EC.visibility_of_element_located((By.ID, "UpdPanel_CU_NMCARD"))).send_keys(namaNasabahAddOn)
        time.sleep(1)
        WebDriverWait(driver, 620).until(EC.visibility_of_element_located((By.ID, "UpdPanel_df_CP_APRLIMIT"))).clear()
        time.sleep(3)
        WebDriverWait(driver, 620).until(EC.visibility_of_element_located((By.ID, "UpdPanel_df_CP_APRLIMIT"))).send_keys(limitData)
        time.sleep(1)
        save_screenshot(driver, '12.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_Button6"))).click()
        time.sleep(3)

        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                logger.info(f"Alert: {alert.text}")
                alert.accept()
                time.sleep(3)
            except:
                pass

        driver.switch_to.default_content()
        logger.info("Verphone Add Basic Step End")

        # ========== DM ADD BASIC ASSIGNMENT ==========
        logger.info("DM Add Basic Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DM%20Add%20Basic%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=6.6.0&tc2=6.6.1&stg=DMBAS")
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, 'MY_BRI_APPROVAL')
        save_screenshot(driver, '13.png')
        logger.info("DM Add Basic Assignment Step End")

        # ========== DATA MAINTENANCE FOR ADD BASIC ==========
        logger.info("Data Maintenance for Add Basic Step Start")
        update_user(valueNoAplikasi)
        logger.info("Update user successful")
        time.sleep(3)

        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DMBAS&passurl&mntitle=Data%20Maintenance%20for%20Add%20Basic&tc=6.6.1&regno={valueNoAplikasi}")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        time.sleep(1)
        save_screenshot(driver, '14.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_btn_appr"))).click()
        time.sleep(3)

        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(3)
            except:
                pass

        driver.switch_to.default_content()
        logger.info("Data Maintenance for Add Basic Step End")

        # ========== DATA MAINTENANCE FOR ADD BASIC APPROVAL ==========
        logger.info("Data Maintenance for Add Basic Approval Step Start")
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DMBAS&passurl&mntitle=Data%20Maintenance%20for%20Add%20Basic%20Approval&tc=6.6.2&regno={valueNoAplikasi}")
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        time.sleep(1)
        save_screenshot(driver, '15.png')

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "UpdPanel_btn_appr"))).click()
        time.sleep(3)

        for _ in range(2):
            try:
                WebDriverWait(driver, 50).until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(3)
            except:
                pass

        driver.switch_to.default_content()
        logger.info("Data Maintenance for Add Basic Approval Step End")

        # ========== COMPLETION ==========
        logger.info("Proses Onboarding Data Add On Basic LOS Selesai")
        elapsed_seconds = int(time.time() - start_time)
        logger.info(f"Proses berjalan selama: {elapsed_seconds} detik")

    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        driver.quit()


if __name__ == "__main__":
    main()