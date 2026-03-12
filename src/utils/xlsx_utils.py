from openpyxl import load_workbook

def read_excel_data(file_path, sheet_name, row, column):
    """
    Perform Read Excel Data with automatic column offset detection.
    
    If column 1 of the sheet has header 'ID', the provided 'column' index 
    is automatically incremented by 1 to maintain compatibility with 
    scripts using hardcoded indices from the old template.
    """
    # Load the Excel workbook
    workbook = load_workbook(file_path)
    # Select the sheet by name
    sheet = workbook[sheet_name]
    
    # Check for 'ID' column at index 1 to handle pergeseran kolom
    first_col_header = sheet.cell(row=1, column=1).value
    if first_col_header == 'ID':
        column += 1
        
    # Retrieve data from specified row and column
    cell_value = sheet.cell(row=row, column=column).value
    return cell_value

def get_row_indices_by_id(sheet, target_id):
    """
    Get all row indices that match the target_id in the first column.
    """
    indices = []
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        cell_id = sheet.cell(row=row, column=1).value
        # Handle cases where ID might be read as int or string
        if cell_id is not None and str(cell_id) == str(target_id):
            indices.append(row)
    return indices

def get_max_row_with_data(sheet):
    """
    Get the last row index that has data in the ID column.
    """
    max_row = sheet.max_row
    for row in range(max_row, 1, -1):
        if sheet.cell(row=row, column=1).value is not None:
            return row
    return 1

def count_filled_rows(sheet):
    """
    Counts rows that have at least one cell filled with data.
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    filled_rows = 0
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if sheet.cell(row=row, column=col).value not in (None, ''):
                filled_rows += 1
                break 
    
    return filled_rows


# # Membuka file Excel
# workbook = load_workbook('Data/LOSData.xlsx')
# sheet = workbook['Card']
# # Contoh penggunaan
# filled_rows = count_filled_rows(sheet)
# print(f"Jumlah baris yang diisi: {filled_rows}")