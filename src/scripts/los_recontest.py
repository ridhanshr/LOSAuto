import os
import sys

# Ensure project root is in sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from src.utils.db_utils import update_appflag, update_user, update_accnum,check_appflag, update_applicant, update_norek
from src.utils.login_utils import perform_login
from src.utils.xlsx_utils import read_excel_data, count_filled_rows
from src.utils.docCheck_utils import update_document
from src.utils.dataEntry_utils import select_dropdown_by_value
from src.utils.assign_utils import assign_data
from src.utils.autoScoring_utils import automate_scoring
from src.utils.ui_utils import setup_webdriver
from src.config import config

import time
import re
import sys
import random
from faker import Faker


def run():
    # Parse command line arguments
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--browser", default=config.get("browser", "edge"), help="Browser type")
    args, unknown = parser.parse_known_args()

    # Setup webdriver based on argument or config
    fast = config.get("fast_mode", False)
    driver = setup_webdriver(args.browser, fast_mode=fast)

    try:
        """
        setup data
        """
        # Setup path data
        xlsxdataPath = r'Data/LOSData.xlsx'
        sheetEntryData = "Entry Data"
        sheetDocCheck = "Document Checking"
        sheetJobInfo = "Job Info"
        sheetAssignment = "Assignment"
        sheetEmergency = "Emergency Contact"
        sheetOtherOptions = "Other Options"
        sheetCard = "Card"

        # Retrieve Data
        # Data Entry Data
        fake = Faker('en_NZ')
        qrCodeData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 1)
        branchData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 2)
        channelData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 3)
        kodeProgramData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 4)
        jenisNasabahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 5)
        salesIDData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 6)
        namaNasabahData = fake.name_male()
        namaPanggilanData = namaNasabahData
        namaCetakArr = namaNasabahData.split(' ')
        namaCetakData = namaCetakArr[0]
        jenisKelaminData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 10)
        tempatLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 12)
        tanggalLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 13)
        bulanLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 14)
        tahunLahirData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 15)
        eKTPData = str(random.randint(10**15, 10**16 - 1))
        statusPernikahanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 17)
        jumlahTanggunganData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 18)
        pendidikanData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 19)
        namaIbuData = fake.name_female()
        identitasPajakData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 21)
        alamatRumahData1 = read_excel_data(xlsxdataPath, sheetEntryData, 2, 22)
        alamatRumahData2 = read_excel_data(xlsxdataPath, sheetEntryData, 2, 23)
        alamatRumahData3 = read_excel_data(xlsxdataPath, sheetEntryData, 2, 24)
        alamatRT = read_excel_data(xlsxdataPath, sheetEntryData, 2, 25)
        alamatRW = read_excel_data(xlsxdataPath, sheetEntryData, 2, 26)
        alamatLurah = read_excel_data(xlsxdataPath, sheetEntryData, 2, 27) 
        alamatCamat = read_excel_data(xlsxdataPath, sheetEntryData, 2, 28) 
        kotaData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 29)
        kodePosData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 30)
        telpRumahPrefixData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 31)
        telpRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 32)
        telpHPData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 33)
        statusRumahData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 34)
        lamaMenempatiThnData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 35)
        lamaMenempatiBlnData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 36)
        NPWPData = str(random.randint(10**14, 10**15 - 1))
        emailData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 38)
        noAplikasiData = read_excel_data(xlsxdataPath, sheetEntryData, 2, 39)

        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=7).value = namaNasabahData
        sheet1.cell(row=2, column=8).value = namaPanggilanData
        sheet1.cell(row=2, column=9).value = namaCetakData
        sheet1.cell(row=2, column=16).value = eKTPData
        sheet1.cell(row=2, column=20).value = namaIbuData
        sheet1.cell(row=2, column=37).value = NPWPData
        workbook.save(xlsxdataPath)

        # Document Check Data
        docCheckData1 = read_excel_data(xlsxdataPath, sheetDocCheck, 2, 1)
        docCheckData2 = read_excel_data(xlsxdataPath, sheetDocCheck, 3, 1)
        docCheckData3 = read_excel_data(xlsxdataPath, sheetDocCheck, 4, 1)
        docCheckData4 = read_excel_data(xlsxdataPath, sheetDocCheck, 5, 1)
        docCheckData6 = read_excel_data(xlsxdataPath, sheetDocCheck, 7, 1)
        docCheckData7 = read_excel_data(xlsxdataPath, sheetDocCheck, 8, 1)

        # Job Info Data
        namaPerusahaanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 1)
        kategoriPekerjaanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 2)
        bidangUsahaData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 3)
        lamaBekerjaThnData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 4)
        lamaBekerjaBlnData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 5)
        pangkatData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 6)
        pendapatanData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 7)
        alamatKantor1 = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 8)
        alamatKantor2 = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 9)
        alamatKantor3 = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 10)
        alamatKantorRT = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 11)
        alamatKantorRW = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 12)
        alamatKantorLurah = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 13)
        alamatKantorCamat = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 14)
        alamatKantorPos = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 15)
        kotaKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 16)
        telpKantorPrefixData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 17)
        telpKantorData = read_excel_data(xlsxdataPath, sheetJobInfo, 2, 18)

        # Emergency Contact Data
        namaEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 1)
        hubunganEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 2)
        alamatEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 3)
        kotaEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 4)
        posEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 5)
        telpRumahPrefixEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 6)
        telpRumahEmergency = read_excel_data(xlsxdataPath, sheetEmergency, 2, 7)

        #Other Options Data
        tagihanData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 1)
        kartuUtamaData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 2)
        pilihanBSData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 3)
        noRekData = read_excel_data(xlsxdataPath, sheetOtherOptions, 2, 4)

        #Card Data
        networkData = read_excel_data(xlsxdataPath, sheetCard, 2, 1)
        productData = read_excel_data(xlsxdataPath, sheetCard, 2, 2)
        cardTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 3)
        plasticTypeData = read_excel_data(xlsxdataPath, sheetCard, 2, 4)
        limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)

        # Assignment Data
        sendToData = read_excel_data(xlsxdataPath, sheetAssignment, 2, 1)
        sendToAppReviewer = read_excel_data(xlsxdataPath, sheetAssignment, 2, 2)

        start_time = time.time()
        """
        login
        """
        # access los
        driver.get("http://172.24.141.61/bricc/unittest.aspx")
        print("Login")
        perform_login(driver)

        """
        generate QR
        """
        time.sleep(2)
        print("Generate QR Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=BAR|G&passurl&mntitle=Generate%20Barcode&row=8&barfontsize=40&barsize=100&imgwidth=250&borderwidth=0&altalign=1&padtop=20&rowheight=73&colwidth=350&cellspacing=0")

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # access element iframe
        driver.find_element(By.XPATH, "//input[@id='mainPanel_branchPanel_BRANCHID_txtID']").send_keys('99010')
        driver.find_element(By.XPATH, "//input[@id='mainPanel_btn_gen']").click()
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        image_element = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='mainPanel_tbl']/tbody/tr[2]/td[1]/img")))
        image_url = image_element.get_attribute('src')
        qrCode = image_url.split("data=")[1].split("&")[0]
        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=1).value = qrCode
        workbook.save(xlsxdataPath)

        # exit from iframe 
        driver.switch_to.default_content()

        print("Generate QR Step End")

        """
        Initial Data Entry
        """
        print("Initial Data Entry Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=IDE|C&passurl&mntitle=Re-Contest&tc=1.0&stg=DE&reap=CON")

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_Button1"))).click()
        time.sleep(1)
        driver.find_element(By.ID, "PopFindBasic_PNL_FindBasic_fAP_REGNO").send_keys(noAplikasiData)
        time.sleep(2)
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.XPATH, "(//input[@value='Find'])[1]"))).click()
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "PopFindBasic_PNL_FindBasic_GridViewX_cell0_7_Button4"))).click()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        driver.find_element(By.XPATH, "(//input[@id='mainPanel_df_BARCODE'])[1]").send_keys('9901000000007')
        time.sleep(3)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_eventPanel_df_EVENTID")))
        selectEvent = Select(driver.find_element(By.ID, "mainPanel_eventPanel_df_EVENTID"))
        selectEvent.select_by_value(str(9999999995))

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, f"mainPanel_df_PLASTICID_{plasticTypeData}"))).click()
        time.sleep(1)

        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_Button2"))).click()
        time.sleep(3)
        WebDriverWait(driver, 50).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        print(alert.text)
        alert.accept()  
        time.sleep(5)
        WebDriverWait(driver, 50).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        print(alert.text)
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

        # Exit from iframe 
        driver.switch_to.default_content()
        print('Initial Data Entry Step End')

        """
        Document Checking
        """
        print("Document Checking Step Start")
        time.sleep(2)
        # ambil no aplikasi
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        elementNoAplikasi = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "UC_GeneralInfo1_l1")))
        valueNoAplikasi = elementNoAplikasi.text

        time.sleep(2)
        workbook = load_workbook(xlsxdataPath)
        sheet1 = workbook['Entry Data']
        sheet1.cell(row=2, column=39).value = valueNoAplikasi
        workbook.save(xlsxdataPath)

        # akses elemen yang ada di dalam iframe
        docCheckDataList = [docCheckData1, docCheckData3, docCheckData4, docCheckData6, docCheckData7]
        for docCheckData in docCheckDataList:
            WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "griddoc_header7_Button4"))).click()
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            update_document(driver, docCheckData)
            WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
            time.sleep(3)

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(3)
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "updPanel_btnUpdate"))).click()

        WebDriverWait(driver, 30).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert_text = alert.text
        alert.accept()
        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(2)
        WebDriverWait(driver, 30).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert_text = alert.text
        alert.accept()

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        # exit from iframe 
        driver.switch_to.default_content()

        print("Document Checking Step End")

        """
        Data Entry Assignment
        """
        print("Data Entry Assignment Step Start")

        # Akses Menu Data Entry Assignment
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV|DE&passurl&mntitle=Data%20Entry%20Assignment&li1=LR|A|SPV|DE|00&li2=LR|A|SPV|DE|01&tc1=3.0&tc2=3.1&atype=DE&stg=DE&xlcode=DE")

        WebDriverWait(driver, 320).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        assign_data(driver, "mainPanel_sendtoPanel_SendTo", valueNoAplikasi, sendToData)

        print("Data Entry Assignment Step End")

        """
        Data Entry
        """
        print("Data Entry Step Start")

        # Update User
        # update_user(valueNoAplikasi)
        # time.sleep(3)

        # Akses Menu Data Entry
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=DE&passurl&mntitle=Data%20Entry&tc=3.1&regno={valueNoAplikasi}&stg=DE")

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Personal Info']"))).click()

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # akses elemen yang ada di dalam iframe
        WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='mainPanel_CU_NICKNM']")))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_NMCARD']").send_keys(namaCetakData)
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_BORNPLACE']").send_keys(tempatLahirData)
        driver.find_element(By.ID, "mainPanel_CU_EMAIL").send_keys(emailData)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_ED_CODE")))
        selectPendidikanTerakhir = Select(driver.find_element(By.ID, "mainPanel_CU_ED_CODE"))
        selectPendidikanTerakhir.select_by_value(str(pendidikanData))
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_TAXIDFLAG")))
        selectIdentitasPajak = Select(driver.find_element(By.ID, "mainPanel_CU_TAXIDFLAG"))
        selectIdentitasPajak.select_by_value(str(identitasPajakData))
        driver.find_element(By.XPATH, "//input[@id='mainPanel_CU_MMNMFIRST']").send_keys(namaIbuData)

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        WebDriverWait(driver, 10).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

        # exit from iframe 
        driver.switch_to.default_content()

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Job Info']"))).click()

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # akses elemen yang ada di dalam iframe
        WebDriverWait(driver, 180).until(EC.visibility_of_element_located((By.ID, "mainPanel_CU_COMPNAME")))
        time.sleep(1)
        driver.find_element(By.XPATH, "//select[@id='mainPanel_CU_OFCTCODE']").send_keys("PT")
        driver.find_element(By.ID, "mainPanel_CU_BUSSINESSDESC").send_keys(bidangUsahaData)
        driver.find_element(By.XPATH, "//select[@id='mainPanel_CU_TOTEMP_CODE']").send_keys("4")
        time.sleep(1)
        driver.find_element(By.ID, "mainPanel_CU_WORKSINCEYY").clear()
        time.sleep(1)
        driver.find_element(By.ID, "mainPanel_CU_WORKSINCEYY").send_keys(lamaBekerjaThnData)
        driver.find_element(By.ID, "mainPanel_CU_WORKSINCEMM").clear()
        time.sleep(1)
        driver.find_element(By.ID, "mainPanel_CU_WORKSINCEMM").send_keys(lamaBekerjaBlnData)
        driver.find_element(By.ID, "mainPanel_CU_INCOME").send_keys(pendapatanData)

        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        WebDriverWait(driver, 999).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()

        # exit from iframe 
        driver.switch_to.default_content()

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Emergency Contact']"))).click()

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # akses elemen yang ada di dalam iframe
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_CU_EMRE_CODE")))
        selectHubuganEmergency = Select(driver.find_element(By.ID, "mainPanel_CU_EMRE_CODE"))
        selectHubuganEmergency.select_by_value(str(hubunganEmergency))
        driver.find_element(By.ID, "mainPanel_CU_EMHMADDR1").send_keys(alamatEmergency)
        driver.find_element(By.ID, "mainPanel_CU_EMHMCITY").send_keys(kotaEmergency)

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        WebDriverWait(driver, 10).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

        # exit from iframe 
        driver.switch_to.default_content()

        WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Other Options']"))).click()
        update_accnum(valueNoAplikasi)
        time.sleep(1)

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # akses elemen yang ada di dalam iframe
        WebDriverWait(driver, 180).until(EC.presence_of_element_located((By.ID, f"mainPanel_BILLING_DLV_TYPE_0")))
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mainPanel_BILLING_STAT_TYPE")))
        selectPilihanBS = Select(driver.find_element(By.ID, "mainPanel_BILLING_STAT_TYPE"))
        selectPilihanBS.select_by_value(str(pilihanBSData))
        driver.find_element(By.ID, "mainPanel_AD_ACCNUM").send_keys('0')

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "mainPanel_BTN_SAVE"))).click()
        time.sleep(3)
        WebDriverWait(driver, 50).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

        # exit from iframe 
        driver.switch_to.default_content()

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Main']"))).click()
        time.sleep(3)

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
        WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        WebDriverWait(driver, 620).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 620).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        WebDriverWait(driver, 620).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))
        print("Data Entry Step End")

        """
        DE Validation Assignment
        """
        print("DE Validation Assignment Step Start")
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=DE%20Validation%20Assignment&li1=LR|SPV|DE|00&li2=LR|SPV|01&tc1=3.5.0&tc2=3.5.1&stg=VDE")

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, sendToData)

        print("DE Validation Assignment Step End")

        """
        DE Validation
        """
        time.sleep(3)
        print("Data Entry Validation Step Start")

        # Update User
        update_user(valueNoAplikasi)
        print("sukses update user")
        time.sleep(3)

        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=VDE&passurl&mntitle=Data%20Entry%20Validation&tc=3.5.1&regno={valueNoAplikasi}&stg=DE")
        time.sleep(3)

        # Access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # Access elemen iframe
        WebDriverWait(driver, 320).until(EC.element_to_be_clickable((By.ID, "mainPanel_btn_update"))).click()
        WebDriverWait(driver, 50).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        print(alert.text)
        alert.accept()  
        time.sleep(5)
        WebDriverWait(driver, 50).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        print(alert.text)
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 180).until(EC.invisibility_of_element_located((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[1]/td[2]')))

        # Exit from iframe 
        driver.switch_to.default_content()

        print("Data Entry Validation Step End")

        """
        Scoring
        """
        host = '172.24.141.33'
        username = 'administrator'
        password = 'c@rdl1nk'
        pathWinSCP = r'C:/Program Files (x86)/WinSCP/WinSCP.exe'

        print("Proses Scoring Mulai")
        appFlagAfter = automate_scoring(pathWinSCP, valueNoAplikasi, host, username, password)

        if appFlagAfter in ["9.2.1", "9.2.3"]:
            print(f"Program dihentikan karena appFlagAfter {appFlagAfter}")
            # Instead of sys.exit, we should probably just return or raise exception
            # But sys.exit is fine if it's running as subprocess
            sys.exit(f"Program dihentikan karena appFlagAfter {appFlagAfter}")
        elif appFlagAfter in ["3.5.0", "3.6.1"]:
            while appFlagAfter in ["3.5.0", "3.6.1"]:
                print(f"Step Scoring Diulang karena flag {appFlagAfter}")
                appFlagAfter = automate_scoring(pathWinSCP, valueNoAplikasi, host, username, password)

        print("Proses Scoring Selesai")

        """
        Credit Reviewer Assignment
        """
        print("Credit Reviewer Assignment Start")

        # Akses Menu Credit Reviewer Assignment
        time.sleep(3)
        driver.get("http://172.24.141.61/bricc/ScreenMenu.aspx?sm=L|SPV&passurl&mntitle=Credit%20Reviewer%20Assignment&li1=LR|SPV|00&li2=LR|SPV|01&tc1=5.7.0&tc2=5.7.1&stg=ANLMINI")
        time.sleep(3)

        WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[2]/tbody[1]/tr[2]/td[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        assign_data(driver, "mainPanel_SendTo", valueNoAplikasi, sendToAppReviewer)

        print("Credit Reviewer Assignment End")

        """
        Credit Reviewer
        """
        print("Credit Reviewer Start")

        # Update User
        update_user(valueNoAplikasi)
        time.sleep(3)

        # Akses Menu Credit Reviewer 
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=ANLMINI&passurl&mntitle=Credit%20Reviewer&tc=5.7.1&regno={valueNoAplikasi}&stg=ANL")

        time.sleep(10)
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Decision']"))).click()
        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        # akses elemen iframe
        limitData = read_excel_data(xlsxdataPath, sheetCard, 2, 5)
        time.sleep(5)
        # akses elemen iframe
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "mainPanel_GV_PRODUCT_cell0_6_BTN_EDIT"))).click()
        time.sleep(3)
        driver.find_element(By.ID, "popupDetail_panelDetail_df_CP_APRLIMIT").send_keys(str(limitData))
        WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.ID, "popupDetail_panelDetail_Button6"))).click()
        time.sleep(10)

        WebDriverWait(driver, 620).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        WebDriverWait(driver, 620).until(EC.invisibility_of_element_located((By.ID, 'popupDetail_CSD-1')))
        time.sleep(3)

        WebDriverWait(driver, 999).until(EC.presence_of_element_located((By.ID, 'mainPanel_decPanel_dfa_AP_NEXTTRBY')))
        select_element = Select(driver.find_element(By.ID, 'mainPanel_decPanel_dfa_AP_NEXTTRBY'))
        select_element.select_by_value("DIRTONI")
        time.sleep(3)

        WebDriverWait(driver, 999).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(3)


        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='mainPanel_decPanel_btn_update']"))).click()
        WebDriverWait(driver, 999).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        time.sleep(3)
        WebDriverWait(driver, 999).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        # exit from iframe 
        driver.switch_to.default_content()

        print("Credit Reviewer  Step End")

        """
        login
        """
        # access los
        driver.get("http://172.24.141.61/bricc/unittest.aspx")
        print("Login")
        perform_login(driver, username='tsipsd')

        """
        Approval
        """
        print("Approval Step Start")
        time.sleep(5)

        # Update User
        update_user(valueNoAplikasi, ap_nexttrby='tsipsd')
        print("update user berhasil")
        time.sleep(5)
        update_user(valueNoAplikasi, ap_nexttrby='tsipsd')
        print("update user berhasil")

        # akses menu approval
        driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=APRV&passurl&mntitle=Approval&tc=7.1&regno={valueNoAplikasi}&stg=APRV")
        time.sleep(10)
        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Decision']"))).click()

        # access iframe
        WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))

        WebDriverWait(driver, 999).until(EC.element_to_be_clickable((By.ID, "mainPanel_decPanel_btn_appr"))).click()
        WebDriverWait(driver, 999).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        WebDriverWait(driver, 999).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))
        time.sleep(5)

        check81 = check_appflag(str(valueNoAplikasi))
        pattern = re.compile(r"'(.*?)'")
        is81 = pattern.search(str(check81))

        if (is81.group(1) == '7.1') :
            while is81.group(1) == '7.1':
                print(f"Proses Approval Diulang karena flag {is81.group(1)}")
                print("Approval Step Start")
                time.sleep(5)
                # akses menu approval
                driver.get(f"http://172.24.141.61/bricc/ScreenMenu.aspx?sm=APRV&passurl&mntitle=Approval&tc=7.1&regno={valueNoAplikasi}&stg=APRV")
                time.sleep(5)

                driver.find_element(By.XPATH, "//a[normalize-space()='Decision']").click()
                # access iframe
                WebDriverWait(driver, 999).until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'framex')))
                WebDriverWait(driver, 620).until(EC.element_to_be_clickable((By.ID, "mainPanel_decPanel_btn_appr"))).click()
                WebDriverWait(driver, 30).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert.accept()
                time.sleep(5)
                WebDriverWait(driver, 180).until(EC.invisibility_of_element((By.XPATH, '//body[1]/form[1]/div[3]/table[4]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/table[2]/tbody[1]/tr[1]/td[2]/span[1]')))

        print("Approval Step End")

        # exit from iframe 
        driver.switch_to.default_content()
        time.sleep(10)

        if (noRekData != 0):
            print("Update Norek Data")
            time.sleep(3)
            update_norek(str(valueNoAplikasi), str(noRekData))
            time.sleep(3)
            print("Update Norek Data Selesai")

        print("Proses Onboarding Data LOS Selesai")
        elapsed_seconds = int(time.time() - start_time)
        print(f"Proses berjalan selama : {elapsed_seconds} detik")

    finally:
        driver.quit()

if __name__ == "__main__":
    run()